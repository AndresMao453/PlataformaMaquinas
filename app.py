# app.py
from __future__ import annotations

import os
import io
import time
import re  # ✅ parse robusto fechas/periodos
import threading
import inspect
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional
import requests

import numpy as np  # ✅ (solicitado)
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from werkzeug.utils import secure_filename

import config
from analyses.registry import get_analyses
import analyses  # noqa: F401
import analyses.continuidad  # ✅ fuerza el register_analysis("continuidad", ...)

# ✅ Loader CSV Google Sheets (por GID)
try:
    from excel_loader import load_gsheet_tab_csv  # cuando excel_loader.py está en raíz
except Exception:
    from services.excel_loader import load_gsheet_tab_csv  # si tu proyecto lo maneja como services/

from analyses.aplicacion_excel import get_period_options_aplicacion
from analyses.aplicacion_excel import run_aplicacion_analysis
from analyses.aplicacion_excel import get_crimpado_export_day_exports

# =========================
# CACHE EN MEMORIA (THB)
# =========================
THB_CACHE: dict[str, dict] = {}

# =========================
# CACHE EN MEMORIA (GSHEET HP)
# =========================
_GSHEET_HP_CACHE: dict[tuple[str, int], tuple[float, pd.DataFrame]] = {}


def load_gsheet_tab_csv_smart(sheet_id: str, gid: int, ttl_s: int) -> pd.DataFrame:
    """
    Lee una pestaña de Google Sheets (CSV) soportando encabezado desplazado.

    IMPORTANTE:
    - No se agrega cache_bust (_=timestamp) al endpoint export, porque Google
      puede responder HTTP 400 en algunos Sheets/GID.
    - Prueba export CSV normal y luego gviz CSV como respaldo.
    - Cachea por (sheet_id, gid) durante ttl_s.
    """
    ttl = int(ttl_s or 10)
    sid = str(sheet_id or "").strip()
    gid_i = int(gid or 0)
    key = (sid, gid_i)
    now = time.time()

    hit = _GSHEET_HP_CACHE.get(key)
    if hit and (now - hit[0]) < ttl:
        return hit[1].copy()

    if not sid or not gid_i:
        raise RuntimeError("Faltan sheet_id/gid para cargar Google Sheets.")

    urls = [
        # ✅ sin cache_bust: evita HTTP 400 en export CSV
        f"https://docs.google.com/spreadsheets/d/{sid}/export?format=csv&gid={gid_i}",
        # ✅ fallback alternativo que suele funcionar cuando export falla
        f"https://docs.google.com/spreadsheets/d/{sid}/gviz/tq?tqx=out:csv&gid={gid_i}",
        f"https://docs.google.com/spreadsheets/d/{sid}/export?format=csv&single=true&gid={gid_i}",
    ]

    last_exc = None
    raw = None
    for url in urls:
        try:
            raw = pd.read_csv(url, header=None)
            break
        except Exception as e:
            last_exc = e

            # segundo intento con requests + User-Agent; ayuda en algunas redes/proxies
            try:
                resp = requests.get(
                    url,
                    timeout=25,
                    headers={"User-Agent": "Mozilla/5.0 PlataformaMaquinas/1.0"},
                )
                resp.raise_for_status()
                raw = pd.read_csv(io.StringIO(resp.text), header=None)
                break
            except Exception as e2:
                last_exc = e2
                continue

    if raw is None:
        raise RuntimeError(f"No se pudo leer Google Sheets CSV sheet_id={sid} gid={gid_i}: {last_exc}")

    raw = raw.dropna(how="all").reset_index(drop=True)

    def _make_unique(cols):
        out = []
        seen = {}
        for i, c in enumerate(cols):
            name = str(c).strip()
            if not name or name.lower() in ("nan", "none"):
                name = f"col_{i + 1}"
            base = name
            k = seen.get(base, 0)
            if k:
                name = f"{base}_{k + 1}"
            seen[base] = k + 1
            out.append(name)
        return out

    best_df = None
    for h in (0, 1, 2, 3, 4, 5):
        if h >= len(raw):
            continue

        cols = _make_unique(raw.iloc[h].tolist())
        df = raw.iloc[h + 1:].copy()
        df.columns = cols

        cols_norm = [str(c).strip().lower() for c in df.columns]
        if any("fecha" in c for c in cols_norm) and any("hora" in c for c in cols_norm):
            best_df = df
            break

    if best_df is None:
        # fallback conservador: primera fila como encabezado
        cols = _make_unique(raw.iloc[0].tolist()) if len(raw) else []
        best_df = raw.iloc[1:].copy() if len(raw) else pd.DataFrame()
        if len(cols) == len(best_df.columns):
            best_df.columns = cols

    _GSHEET_HP_CACHE[key] = (now, best_df.copy())
    return best_df.copy()


# =========================
# CACHE EXCELFILE + PARSE (GENERAL)
# =========================
CACHE_TTL_SEC = 45 * 60  # 45 min
MAX_FILES_IN_RAM = 10  # límite de ExcelFiles en RAM
MAX_PARSES_PER_FILE = 50  # límite de hojas/headers cacheados por archivo

_excel_lock = threading.RLock()


def _normalize_gsheet_paradas_duration_to_seconds(df_paradas: "pd.DataFrame") -> "pd.DataFrame":
    """
    Normaliza Duración/tiempo en Paradas cuando viene de Google Sheets (CSV).
    Convierte a SEGUNDOS (float) sin tocar el análisis.

    Soporta:
    - "HH:MM" / "HH:MM:SS"
    - "5 min", "5 minutos", "30 s", "30 seg"
    - numérico:
        * <=2  -> fracción de día (Excel) => *86400
        * valores típicos (1..240) -> MINUTOS => *60
        * si no, se deja como segundos
    """
    import pandas as pd
    import re

    if df_paradas is None or not isinstance(df_paradas, pd.DataFrame) or df_paradas.empty:
        return df_paradas

    def _norm(x):
        return str(x or "").strip().lower()

    # 1) elegir mejor columna de duración
    dur_col = None
    best_score = -1

    for c in df_paradas.columns:
        k = _norm(c)
        if ("dur" in k) or ("tiempo" in k):
            s = df_paradas[c].astype(str).str.replace("\u00a0", " ", regex=False).str.strip().str.lower()
            # score = cuántos parecen tiempo/número
            score = int(s.str.contains(":", regex=False).sum()) + int(
                pd.to_numeric(s.str.replace(",", ".", regex=False), errors="coerce").fillna(0).gt(0).sum())
            if score > best_score:
                best_score = score
                dur_col = c

    if not dur_col:
        return df_paradas

    s = df_paradas[dur_col].astype(str).str.replace("\u00a0", " ", regex=False).str.strip()
    sl = s.str.lower()

    # 2) parse HH:MM / HH:MM:SS
    is_time = sl.str.contains(":", regex=False)
    tsec = pd.to_timedelta(s.where(is_time), errors="coerce").dt.total_seconds()

    # 3) parse con sufijos ("min", "seg")
    #    ejemplo: "5 min", "10 minutos", "30 s", "45seg"
    def _extract_number(txt):
        txt = str(txt).strip().lower().replace(",", ".")
        m = re.search(r"(\d+(?:\.\d+)?)", txt)
        return float(m.group(1)) if m else None

    sec = pd.Series(0.0, index=df_paradas.index, dtype="float64")

    # set time-like seconds
    if tsec is not None:
        sec.loc[is_time] = pd.to_numeric(tsec.loc[is_time], errors="coerce").fillna(0.0)

    # no-time rows
    idx = (~is_time)

    if idx.any():
        vals = s.loc[idx].astype(str)

        # detectar sufijos
        is_min = vals.str.lower().str.contains(r"\bmin\b|\bmins\b|minuto", regex=True, na=False)
        is_sec = vals.str.lower().str.contains(r"\bseg\b|\bsecs\b|\bsec\b|\bs\b", regex=True, na=False)

        # números crudos
        nums = pd.to_numeric(vals.str.replace(",", ".", regex=False), errors="coerce")

        # si tiene sufijo min/seg y no parsea directo, extraer con regex
        need_extract = nums.isna() & (is_min | is_sec)
        if need_extract.any():
            ex = vals.loc[need_extract].apply(_extract_number)
            nums.loc[need_extract] = pd.to_numeric(ex, errors="coerce")

        nz = nums[nums > 0].dropna()

        # default: segundos
        out = nums.copy()

        if not nz.empty:
            mx = float(nz.max())
            med = float(nz.median())

            # fracción de día Excel
            if mx <= 2.0:
                out = nums * 86400.0

            else:
                # si columna/valores parecen MINUTOS (muy común en Sheets)
                # heurística: la mayoría <=240 y med <=60
                looks_minutes = (mx <= 240.0 and med <= 60.0)

                if looks_minutes:
                    out = nums * 60.0

        # forzar por sufijo
        out.loc[is_min] = pd.to_numeric(nums.loc[is_min], errors="coerce").fillna(0.0) * 60.0
        out.loc[is_sec] = pd.to_numeric(nums.loc[is_sec], errors="coerce").fillna(0.0)

        sec.loc[idx] = pd.to_numeric(out, errors="coerce").fillna(0.0)

    df2 = df_paradas.copy()
    df2[dur_col] = sec
    return df2


@dataclass
class _ExcelEntry:
    path: str
    mtime: float
    last_access: float
    xl: pd.ExcelFile
    parses: dict[tuple, pd.DataFrame] = field(default_factory=dict)


_EXCEL_CACHE: dict[str, _ExcelEntry] = {}


def _purge_excel_cache(now: float) -> None:
    # TTL
    expired = [k for k, e in _EXCEL_CACHE.items() if (now - e.last_access) > CACHE_TTL_SEC]
    for k in expired:
        try:
            _EXCEL_CACHE[k].xl.close()
        except Exception:
            pass
        _EXCEL_CACHE.pop(k, None)

    # LRU
    if len(_EXCEL_CACHE) > MAX_FILES_IN_RAM:
        items = sorted(_EXCEL_CACHE.items(), key=lambda kv: kv[1].last_access)  # más viejo primero
        to_remove = len(_EXCEL_CACHE) - MAX_FILES_IN_RAM
        for k, e in items[:to_remove]:
            try:
                e.xl.close()
            except Exception:
                pass
            _EXCEL_CACHE.pop(k, None)


def _get_excel_entry(path: str) -> _ExcelEntry:
    ap = os.path.abspath(path)
    if not os.path.exists(ap):
        raise FileNotFoundError(ap)

    now = time.time()
    mtime = os.path.getmtime(ap)

    with _excel_lock:
        _purge_excel_cache(now)

        e = _EXCEL_CACHE.get(ap)
        if e and e.mtime == mtime:
            e.last_access = now
            return e

        # recargar
        try:
            if e:
                try:
                    e.xl.close()
                except Exception:
                    pass
        except Exception:
            pass

        xl = pd.ExcelFile(ap, engine="openpyxl")
        e2 = _ExcelEntry(path=ap, mtime=mtime, last_access=now, xl=xl)
        _EXCEL_CACHE[ap] = e2
        return e2


def read_excel_cached(path: str, sheet_name: str, header=0, usecols=None, copy: bool = False) -> pd.DataFrame:
    """
    Cachea parse por (sheet_name, header, usecols). Devuelve DF (opcionalmente copia).
    """
    e = _get_excel_entry(path)
    key = (sheet_name, header, usecols)

    with _excel_lock:
        df = e.parses.get(key)
        if df is not None:
            e.last_access = time.time()
            return df.copy() if copy else df

    df = e.xl.parse(sheet_name=sheet_name, header=header, usecols=usecols)

    with _excel_lock:
        # control simple de tamaño del dict de parses
        if len(e.parses) >= MAX_PARSES_PER_FILE:
            # purga LRU de parses (por simplicidad: borrar el primero)
            try:
                e.parses.pop(next(iter(e.parses)))
            except Exception:
                e.parses = {}
        e.parses[key] = df
        e.last_access = time.time()

    return df.copy() if copy else df


def warm_excel_cache(path: str) -> None:
    """Abre ExcelFile una vez (warm-up)."""
    try:
        _get_excel_entry(path)
    except Exception:
        pass


# =========================
# LECTURA SMART (USANDO CACHE)
# =========================
def _read_excel_smart(path: str, sheet_name: str) -> pd.DataFrame:
    """
    Lee una hoja probando header=0/1/2 y se queda con la que tenga columnas útiles.
    HPAndres.xlsx usa header=1 (fila 2) en sus 3 hojas.
    """
    for h in (0, 1, 2):
        df = read_excel_cached(path, sheet_name=sheet_name, header=h, copy=False)
        cols = [str(c).strip().lower() for c in df.columns]
        if any("fecha" in c for c in cols) and any("hora" in c for c in cols):
            return df
    # fallback
    return read_excel_cached(path, sheet_name=sheet_name, header=0, copy=False)


def _find_col(df: pd.DataFrame, needle: str) -> str | None:
    n = needle.strip().lower()
    for c in df.columns:
        if n in str(c).strip().lower():
            return c
    return None


def _load_all_aplicacion_4maq(ttl_s: int | None = None) -> pd.DataFrame:
    """
    Une en un solo DF los 4 registros de APLICACIÓN (máquinas).
    Usa tu config.py: GSHEET_ID..GSHEET_ID4 y GID_APLICACION..GID_APLICACION4
    """
    # OJO: en tu proyecto real esto suele estar en services/excel_loader.py
    from services.excel_loader import load_gsheet_tab_csv

    ttl = int(getattr(config, "GSHEET_TTL_S", 10) or 10) if ttl_s is None else int(ttl_s)

    sources = [
        ("M1", getattr(config, "GSHEET_ID", None), getattr(config, "GSHEET_GID_APLICACION", None)),
        ("M2", getattr(config, "GSHEET_ID2", None), getattr(config, "GSHEET_GID_APLICACION2", None)),
        ("M3", getattr(config, "GSHEET_ID3", None), getattr(config, "GSHEET_GID_APLICACION3", None)),
        ("M4", getattr(config, "GSHEET_ID4", None), getattr(config, "GSHEET_GID_APLICACION4", None)),
    ]

    frames = []
    for tag, sid, gid in sources:
        if not sid or gid is None:
            continue
        df = load_gsheet_tab_csv(str(sid), int(gid), ttl_s=ttl)
        df["__src_machine"] = tag
        frames.append(df)

    if not frames:
        return pd.DataFrame()

    out = pd.concat(frames, ignore_index=True)
    return out


def _build_dt_from_fecha_hora(df, col_fecha, col_hora):
    dt_fecha = _coerce_excel_date_series(df[col_fecha])  # <- tu helper robusto (el de app.py)
    if not col_hora or col_hora not in df.columns:
        return dt_fecha

    sec = _excel_time_to_seconds(df[col_hora]).fillna(0)
    return dt_fecha + pd.to_timedelta(sec, unit="s")


def _hp_gsheet_fix_paradas_duration(df_paradas: "pd.DataFrame") -> "pd.DataFrame":
    """
    HP (Google Sheets CSV): Duración suele venir en MINUTOS (1,5,10,30...)
    pero el análisis la asume en SEGUNDOS si es numérica.
    Este fix detecta eso y convierte a segundos.
    """
    import pandas as pd

    if df_paradas is None or not isinstance(df_paradas, pd.DataFrame) or df_paradas.empty:
        return df_paradas

    # detectar columna "duración" candidata (sin tocar fecha/hora/código)
    def _n(x):
        return str(x or "").strip().lower()

    cols = list(df_paradas.columns)

    # excluir columnas típicas
    ex = set()
    for c in cols:
        k = _n(c)
        if ("fecha" in k) or ("hora" in k) or ("código" in k) or ("codigo" in k):
            ex.add(c)

    cand = []
    for c in cols:
        if c in ex:
            continue
        k = _n(c)
        if ("dur" in k) or ("tiempo" in k) or ("min" in k) or ("seg" in k) or ("sec" in k):
            cand.append(c)

    if not cand:
        return df_paradas

    # escoger la que más números >0 tenga
    best = None
    best_nz = -1
    best_num = None

    for c in cand:
        s = df_paradas[c].astype(str).str.replace("\u00a0", " ", regex=False).str.strip()
        # si viene con ":" probablemente ya es tiempo (HH:MM/HH:MM:SS) -> no tocar
        if s.str.contains(":", regex=False).mean() > 0.2:
            continue

        num = pd.to_numeric(s.str.replace(",", ".", regex=False), errors="coerce")
        nz = int((num.fillna(0) > 0).sum())
        if nz > best_nz:
            best_nz = nz
            best = c
            best_num = num

    if not best or best_nz <= 0:
        return df_paradas

    name = _n(best)
    nz = best_num[best_num > 0].dropna()
    if nz.empty:
        return df_paradas

    # -------------------------
    # decidir unidad
    # -------------------------
    # 1) Si header dice minutos => minutos
    if ("min" in name) and not (("seg" in name) or ("sec" in name)):
        is_minutes = True
    # 2) Si header dice segundos => segundos
    elif ("seg" in name) or ("sec" in name):
        is_minutes = False
    else:
        # 3) Heurística: si casi todo es <60 (y por eso se ve 00:00), probablemente son minutos.
        pct_lt60 = float((nz < 60).mean())
        mx = float(nz.max())
        med = float(nz.median())
        # típico en minutos: valores pequeños (1..60) y muchísimos <60
        is_minutes = (pct_lt60 >= 0.75 and mx <= 240 and med <= 60)

    df2 = df_paradas.copy()
    if is_minutes:
        df2[best] = pd.to_numeric(best_num, errors="coerce").fillna(0.0) * 60.0
    else:
        df2[best] = pd.to_numeric(best_num, errors="coerce").fillna(0.0)

    return df2


# =========================
# HELPERS GENERALES
# =========================
def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in config.ALLOWED_EXTENSIONS


def safe_upload_path(upload_folder: str, filename: str) -> str:
    filename = os.path.basename(filename)
    return os.path.join(upload_folder, filename)


# =========================
# PARSE HELPERS (THB)
# =========================
def _excel_serial_to_datetime(s: pd.Series) -> pd.Series:
    # Excel date serial -> datetime
    return pd.to_datetime(s, unit="D", origin="1899-12-30", errors="coerce")


def _coerce_excel_date_series(s: pd.Series) -> pd.Series:
    """
    Fecha robusta (mixta):
    - datetime/date ya parseado
    - serial Excel numérico
    - serial Excel como string ("45234")
    - string ISO "yyyy-mm-dd"
    - string "dd/mm/yyyy" (Colombia) -> dayfirst=True por defecto
      (solo monthfirst cuando el 2do número > 12, ej 01/23/2026)
    Retorna datetime normalizado (00:00:00).
    """
    if s is None:
        return pd.Series([], dtype="datetime64[ns]")

    # ya datetime
    if pd.api.types.is_datetime64_any_dtype(s):
        return pd.to_datetime(s, errors="coerce").dt.normalize()

    # numérico -> serial Excel
    if pd.api.types.is_numeric_dtype(s):
        dt0 = pd.to_datetime(s, unit="D", origin="1899-12-30", errors="coerce")
        return pd.to_datetime(dt0, errors="coerce").dt.normalize()

    # string/object
    txt = s.astype(str).str.strip()

    # vacíos a NaN
    txt = txt.replace({"": None, "nan": None, "None": None, "NaT": None})

    out = pd.Series(pd.NaT, index=s.index, dtype="datetime64[ns]")

    # 1) serial Excel como string (sin / ni -)
    num = pd.to_numeric(txt, errors="coerce")
    mask_serial = num.notna() & ~txt.str.contains(r"[/-]", regex=True, na=False)
    if mask_serial.any():
        out.loc[mask_serial] = pd.to_datetime(
            num.loc[mask_serial],
            unit="D",
            origin="1899-12-30",
            errors="coerce",
        )

    # 2) ISO yyyy-mm-dd o yyyy/mm/dd
    mask_iso = txt.str.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}$", na=False)
    if mask_iso.any():
        out.loc[mask_iso] = pd.to_datetime(
            txt.loc[mask_iso],
            errors="coerce",
            yearfirst=True,
            dayfirst=False,
        )

    # 3) dd/mm/yyyy o mm/dd/yyyy (ambigua) -> resolver con regla
    mask_dmy = txt.str.match(r"^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$", na=False) & ~mask_iso
    if mask_dmy.any():
        # parse ambos sentidos
        dt_day = pd.to_datetime(txt, errors="coerce", dayfirst=True)
        dt_mon = pd.to_datetime(txt, errors="coerce", dayfirst=False)

        # extraer a/b para decidir
        t2 = txt.str.replace("-", "/", regex=False)
        parts = t2.str.extract(r"^(?P<a>\d{1,2})/(?P<b>\d{1,2})/(?P<y>\d{2,4})$")
        a = pd.to_numeric(parts["a"], errors="coerce")
        b = pd.to_numeric(parts["b"], errors="coerce")

        # si b>12 y a<=12 -> es mm/dd (monthfirst)
        mask_mdy = mask_dmy & (b > 12) & (a <= 12)

        out.loc[mask_mdy] = dt_mon.loc[mask_mdy]
        out.loc[mask_dmy & ~mask_mdy] = dt_day.loc[mask_dmy & ~mask_mdy]

    # 4) fallback: lo que falte
    mask_rem = out.isna()
    if mask_rem.any():
        out.loc[mask_rem] = pd.to_datetime(txt.loc[mask_rem], errors="coerce")

    return pd.to_datetime(out, errors="coerce").dt.normalize()


def _excel_time_to_seconds(s: pd.Series) -> pd.Series:
    """
    Convierte Hora a segundos desde 00:00.
    - numérico: fracción de día Excel
    - string: "HH:MM" o "HH:MM:SS"
    """
    if s is None:
        return pd.Series([], dtype="float64")

    if pd.api.types.is_numeric_dtype(s):
        td = pd.to_timedelta(s, unit="D", errors="coerce")
        return td.dt.total_seconds()

    ss = s.astype(str).str.strip()
    td = pd.to_timedelta(ss, errors="coerce")
    return td.dt.total_seconds()


def _seconds_to_hhmm(sec: float) -> str:
    if pd.isna(sec):
        return ""
    sec = int(max(0, sec))
    hh = sec // 3600
    mm = (sec % 3600) // 60
    return f"{hh:02d}:{mm:02d}"


def _parse_hhmm_to_seconds(hhmm: str) -> int:
    hhmm = (hhmm or "").strip()
    if not hhmm:
        return 0
    parts = hhmm.split(":")
    if len(parts) < 2:
        return 0
    hh = int(parts[0])
    mm = int(parts[1])
    return hh * 3600 + mm * 60


def _pick_best_col(df: pd.DataFrame, target: str) -> str | None:
    """
    ✅ FIX solicitado: escoger mejor columna por match exacto y luego 'contiene'
    para evitar elegir una 'Fecha...' rara en vez de 'Fecha' real.
    """
    cols = list(df.columns)
    low = {str(c).strip().lower(): c for c in cols}

    t = target.strip().lower()
    if t in low:
        return low[t]
    for k, orig in low.items():
        if t in k:
            return orig
    return None


# =========================
# CARGA HOJAS THB (USANDO CACHE)
# =========================
def load_thb_verificacion(path: str) -> pd.DataFrame:
    """
    THB: hoja 'Verificación de Medidas'
    Encabezado está en la fila 2 -> header=1
    """
    sheet = "Verificación de Medidas"
    df = read_excel_cached(path, sheet_name=sheet, header=1, copy=True)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _find_header_row(raw: pd.DataFrame, required: set[str], max_scan: int = 30) -> int:
    """
    ✅ FIX solicitado: Paradas a veces trae título arriba, header real más abajo.
    Busca una fila que contenga (al menos) los headers requeridos.
    """
    req = {str(x).strip().lower() for x in required}
    max_i = min(max_scan, len(raw))
    for i in range(max_i):
        row = raw.iloc[i].astype(str).str.strip().str.lower().tolist()
        row_set = set(row)
        if req.issubset(row_set):
            return i
    return 0


def load_thb_paradas(path: str) -> pd.DataFrame:
    """
    ✅ THB hoja 'Paradas' con header embebido.
    - Detecta header real.
    - Detecta descripción.
    - Limpia 'nan/none'.
    - Rellena descripciones vacías por código.
    """
    sheet = "Paradas"
    raw = read_excel_cached(path, sheet_name=sheet, header=None, copy=True)

    hi = _find_header_row(raw, {"Fecha", "Hora"})
    header = raw.iloc[hi].astype(str).str.strip().tolist()

    dfp = raw.iloc[hi + 1:].copy()
    dfp.columns = header
    dfp = dfp.dropna(how="all").reset_index(drop=True)

    dfp.columns = [str(c).strip() for c in dfp.columns]

    # coerce Fecha si existe (robusto)
    fecha_col = _pick_best_col(dfp, "fecha")
    if fecha_col and fecha_col in dfp.columns:
        dfp[fecha_col] = _coerce_excel_date_series(dfp[fecha_col])

    # asegurar "Descripción" usable
    codigo_col = _pick_best_col(dfp, "código") or _pick_best_col(dfp, "codigo")
    desc_col = (
            _pick_best_col(dfp, "descrip")
            or _pick_best_col(dfp, "descripción")
            or _pick_best_col(dfp, "descripcion")
            or _pick_best_col(dfp, "motivo")
    )

    if desc_col and desc_col in dfp.columns:
        s = dfp[desc_col].astype(str).str.strip()
        s = s.where(~s.str.lower().isin(["nan", "none"]), "")
        dfp[desc_col] = s

        if codigo_col and codigo_col in dfp.columns:
            dfp[desc_col] = (
                dfp.groupby(codigo_col, dropna=False)[desc_col]
                .transform(lambda x: x.replace("", pd.NA).ffill().bfill().fillna(""))
            )

    return dfp


def get_thb_datetime_and_seconds(df: pd.DataFrame) -> tuple[pd.Series, pd.Series, str | None, str | None]:
    """
    Retorna:
      - dt_fecha (datetime normalizado)
      - sec_hora (float segundos)
      - fecha_col (nombre real)
      - hora_col (nombre real)
    """
    if df is None or df.empty:
        return pd.Series([], dtype="datetime64[ns]"), pd.Series([], dtype="float64"), None, None

    fecha_col = _pick_best_col(df, "fecha")
    hora_col = _pick_best_col(df, "hora")

    if fecha_col is None:
        return pd.Series([], dtype="datetime64[ns]"), pd.Series([], dtype="float64"), None, None

    dt_fecha = _coerce_excel_date_series(df[fecha_col])

    sec_hora = pd.Series([], dtype="float64")
    if hora_col is not None and hora_col in df.columns:
        sec_hora = _excel_time_to_seconds(df[hora_col])

    return dt_fecha, sec_hora, fecha_col, hora_col


def filter_by_period(dt_fecha: pd.Series, period: str, period_value: str) -> pd.Series:
    """
    ✅ FIX CRÍTICO (THB): el día NO podía filtrar porque había un `pass`.
    Ahora day/week/month filtran de verdad.
    """
    dt = _coerce_excel_date_series(pd.Series(dt_fecha))

    if dt is None or getattr(dt, "empty", False):
        return pd.Series([], dtype="bool")

    if period == "day":
        d = pd.to_datetime(period_value, errors="coerce")
        if pd.isna(d):
            return pd.Series(False, index=dt.index)
        d = pd.Timestamp(d).normalize()
        return dt.dt.normalize().eq(d).fillna(False)

    if period == "week":
        try:
            y, w = str(period_value).replace("w", "W").split("-W")
            y = int(y)
            w = int(w)
        except Exception:
            return pd.Series([False] * len(dt), index=dt.index)

        iso = dt.dt.isocalendar()
        mask = (iso["year"] == y) & (iso["week"] == w)
        return mask.fillna(False)

    if period == "month":
        m = re.match(r"^(\d{4})-(\d{2})$", str(period_value))
        if not m:
            return pd.Series([False] * len(dt), index=dt.index)
        yy = int(m.group(1))
        mm = int(m.group(2))
        mask = (dt.dt.year == yy) & (dt.dt.month == mm)
        return mask.fillna(False)

    return pd.Series([True] * len(dt), index=dt.index)


def build_period_options(dt: pd.Series, period: str) -> list[dict]:
    """
    Opciones day/week/month robustas con fechas coerced.
    """
    if dt is None or getattr(dt, "empty", False):
        return []

    dt = _coerce_excel_date_series(pd.Series(dt)).dropna()
    if dt.empty:
        return []

    if period == "day":
        days = pd.Series(dt.dt.date).dropna().unique()
        days_sorted = sorted(days)
        return [{"value": str(d), "label": d.strftime("%d/%m/%Y")} for d in days_sorted]

    if period == "week":
        iso = dt.dt.isocalendar()
        pairs = sorted(set(zip(iso["year"].tolist(), iso["week"].tolist())))
        return [{"value": f"{int(y)}-W{int(w):02d}", "label": f"Semana {int(w):02d} - {int(y)}"} for y, w in pairs]

    if period == "month":
        months = sorted(set((int(y), int(m)) for y, m in zip(dt.dt.year.tolist(), dt.dt.month.tolist())))
        out = []
        for y, m in months:
            out.append({"value": f"{y}-{m:02d}", "label": f"{m:02d}/{y}"})
        return out

    return []


def build_thb_cache(path: str) -> dict:
    """
    Lee una vez el Excel THB y construye:
      - df verificación
      - df paradas
      - opciones day/week/month
      - para cada (period, value): operarias + min/max hora
    """
    df = load_thb_verificacion(path)

    try:
        df_paradas = load_thb_paradas(path)
    except Exception:
        df_paradas = pd.DataFrame()

    name_col = next((c for c in df.columns if "nombre" in str(c).strip().lower()), None)
    dt_fecha, _, _, hora_col = get_thb_datetime_and_seconds(df)

    cache = {
        "df": df,
        "df_paradas": df_paradas,
        "days": [],
        "weeks": [],
        "months": [],
        "by_period": {"day": {}, "week": {}, "month": {}},
    }

    cache["days"] = build_period_options(dt_fecha, "day")
    cache["weeks"] = build_period_options(dt_fecha, "week")
    cache["months"] = build_period_options(dt_fecha, "month")

    def store(period: str, value: str, mask: pd.Series):
        sub = df.loc[mask].copy()

        ops = []
        if name_col and name_col in sub.columns:
            ops = sorted(sub[name_col].dropna().astype(str).str.strip().unique().tolist())

        min_time = ""
        max_time = ""
        if hora_col and hora_col in sub.columns:
            sec = _excel_time_to_seconds(sub[hora_col]).dropna()
            if not sec.empty:
                min_time = _seconds_to_hhmm(sec.min())
                max_time = _seconds_to_hhmm(sec.max())

        cache["by_period"][period][value] = {"operators": ops, "min_time": min_time, "max_time": max_time}

    for o in cache["days"]:
        v = o["value"]
        store("day", v, filter_by_period(dt_fecha, "day", v))

    for o in cache["weeks"]:
        v = o["value"]
        store("week", v, filter_by_period(dt_fecha, "week", v))

    for o in cache["months"]:
        v = o["value"]
        store("month", v, filter_by_period(dt_fecha, "month", v))

    return cache


# =========================
# FLASK APP
# =========================
app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = config.UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = config.MAX_CONTENT_LENGTH
app.secret_key = config.SECRET_KEY

os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

LINE_UI = {
    "corte": {"title": "Modulo OEE FASECOL (Corte)", "desc": "KPIs, productividad, tiempos y validaciones."},
    "aplicacion": {"title": "Modulo OEE FASECOL (Aplicación)",
                   "desc": "Crimpados, paradas y comportamiento operativo."},
    "continuidad": {"title": "Modulo OEE FASECOL (Continuidad)",
                    "desc": "Resultados de prueba, tasas de OK/NO OK y tiempos."},
    "ciclo_aplicadores": {"title": "Ciclo / Aplicadores",
                          "desc": "Ciclos de máquina, cambios de aplicador y métricas operativas.",
                          },

}

# ENABLED_LINES = {"corte", "aplicacion"}
GSHEET_APP_TOKEN = "__GSHEET_APP__"  # marcador interno (no es un archivo real)
GSHEET_HP_TOKEN = "__GSHEET_HP__"  # ✅ HP por Google Sheets (no es un archivo real)
GSHEET_THB_TOKEN = "__GSHEET_THB__"  # ✅ THB por Google Sheets
ENABLED_LINES = {"corte", "aplicacion", "continuidad", "ciclo_aplicadores"}
GSHEET_CONT_BUSES_TOKEN = "__GSHEET_CONT_BUSES__"
GSHEET_CONT_MOTOS_TOKEN = "__GSHEET_CONT_MOTOS__"


@app.route("/", methods=["GET"])
def home():
    return render_template("home.html")


@app.route("/pantalla_rotativa", methods=["GET"])
def pantalla_rotativa():
    """Pantalla tipo TV: rota automáticamente entre THB y Crimpado Aplicación M1."""
    return render_template("pantalla_rotativa.html")


@app.route("/line/<line_key>", methods=["GET", "POST"])
def line_start(line_key: str):
    analyses_map = get_analyses()
    if line_key not in analyses_map:
        flash("Línea no disponible todavía.", "error")
        return redirect(url_for("home"))

    if line_key not in ENABLED_LINES:
        flash("🚧 Esta interfaz aún no está disponible. Próximamente.", "error")
        return redirect(url_for("home"))

    ui = LINE_UI.get(line_key, {"title": "📊 Análisis", "desc": ""})

    # =========================
    # ✅ GET
    # =========================
    if request.method == "GET":
        # =====================================================
        # ✅ MODO TV / PANTALLA ROTATIVA
        # =====================================================
        # La pantalla rotativa entra por URL directa, por ejemplo:
        #   /line/corte?auto=tv&machine=THB
        #   /line/aplicacion?auto=tv&subcat=aplicacion&machine_id=1
        #   /line/aplicacion?auto=tv&subcat=aplicacion&machine_id=2
        #   /line/aplicacion?auto=tv&subcat=union&machine_id=1
        #   /line/aplicacion?auto=tv&subcat=union&machine_id=2
        #
        # Antes el GET ignoraba esos parámetros y siempre abría stage="upload".
        # En modo TV eso podía dejar la vista en selección y no llegaba a
        # Productividad por hora. Aquí solo se salta esa selección manual para TV.
        auto_raw = (request.args.get("auto") or request.args.get("tv") or "").strip().lower()
        is_tv_auto = auto_raw in ("tv", "1", "true", "yes", "si", "sí")

        if is_tv_auto:
            # -------------------------
            # CORTE TV: THB / HP
            # -------------------------
            if line_key == "corte":
                tv_machine = (request.args.get("machine") or "THB").strip().upper()
                if tv_machine not in ("THB", "HP"):
                    tv_machine = "THB"

                tv_filename = GSHEET_HP_TOKEN if tv_machine == "HP" else GSHEET_THB_TOKEN

                return render_template(
                    "line_start.html",
                    line_key=line_key,
                    line_title=ui["title"],
                    line_desc=ui["desc"],
                    stage="select_sheet",
                    filename=tv_filename,
                    machine=tv_machine,
                    subcat="",
                    machine_id="",
                )

            # -------------------------
            # APLICACIÓN / UNIÓN TV
            # -------------------------
            if line_key == "aplicacion":
                tv_subcat = (request.args.get("subcat") or "aplicacion").strip().lower()
                if tv_subcat not in ("aplicacion", "union"):
                    tv_subcat = "aplicacion"

                tv_machine_id = (request.args.get("machine_id") or "1").strip()
                if tv_machine_id not in ("1", "2"):
                    tv_machine_id = "1"

                tv_machine = "APLICACION" if tv_subcat == "aplicacion" else "UNION"

                return render_template(
                    "line_start.html",
                    line_key=line_key,
                    line_title=ui["title"],
                    line_desc=ui["desc"],
                    stage="select_sheet",
                    filename=GSHEET_APP_TOKEN,
                    machine=tv_machine,
                    subcat=tv_subcat,
                    machine_id=tv_machine_id,
                )

        # ✅ Ciclo/Aplicadores: entra directo (sin upload)
        if line_key == "ciclo_aplicadores":
            return render_template(
                "line_start.html",
                line_key=line_key,
                line_title=ui["title"],
                line_desc=ui["desc"],
                stage="select_sheet",
                filename=GSHEET_APP_TOKEN,
                machine="APLICACION",
                subcat="",
                machine_id="",
            )

        return render_template(
            "line_start.html",
            line_key=line_key,
            line_title=ui["title"],
            line_desc=ui["desc"],
            stage="upload"
        )

    # =========================
    # ✅ POST (validaciones)
    # =========================
    machine = (request.form.get("machine") or "").strip().upper()
    subcat = (request.form.get("subcat") or "").strip().lower()  # solo aplica para aplicacion
    machine_id = (request.form.get("machine_id") or "").strip()  # ✅ Maquina 1/2 para Aplicación/Unión
    source = (request.form.get("source") or "").strip().lower()  # ✅ gsheet / excel

    # -------------------------
    # ✅ CORTE
    # -------------------------
    if line_key == "corte":
        if machine not in ("HP", "THB"):
            flash("Debes seleccionar la máquina (HP o THB) antes de subir el Excel.", "error")
            return redirect(url_for("line_start", line_key=line_key))

        # ✅ CORTE por Google Sheets (sin archivo)
        if source == "gsheet":
            filename = GSHEET_HP_TOKEN if machine == "HP" else GSHEET_THB_TOKEN
            return render_template(
                "line_start.html",
                line_key=line_key,
                line_title=ui["title"],
                line_desc=ui["desc"],
                stage="select_sheet",
                filename=filename,
                machine=machine,
                subcat=subcat,
                machine_id=machine_id,
            )

        # ✅ CORTE por Excel (exige archivo)
        f = request.files.get("file")
        if not f or f.filename == "":
            flash("Selecciona un archivo Excel.", "error")
            return redirect(url_for("line_start", line_key=line_key))

        if not allowed_file(f.filename):
            flash("Formato no permitido. Sube .xlsx o .xls", "error")
            return redirect(url_for("line_start", line_key=line_key))

        safe_name = secure_filename(f.filename)
        filename = f"{int(time.time())}_{safe_name}"
        save_path = safe_upload_path(app.config["UPLOAD_FOLDER"], filename)
        f.save(save_path)

        warm_excel_cache(save_path)

        if machine == "THB":
            try:
                THB_CACHE[filename] = build_thb_cache(save_path)
            except Exception:
                THB_CACHE.pop(filename, None)
                flash("No se pudo construir el índice THB (revisa hoja/columnas).", "error")

        return render_template(
            "line_start.html",
            line_key=line_key,
            line_title=ui["title"],
            line_desc=ui["desc"],
            stage="select_sheet",
            filename=filename,
            machine=machine,
            subcat=subcat,
            machine_id=machine_id,
        )

    # -------------------------
    # ✅ APLICACION / UNION (SIEMPRE Google Sheets)
    # -------------------------
    if line_key in ("aplicacion", "aplicacion_excel"):
        if subcat not in ("aplicacion", "union"):
            flash("Debes seleccionar la categoría (Aplicación o Unión).", "error")
            return redirect(url_for("line_start", line_key=line_key))

        if machine not in ("APLICACION", "UNION"):
            machine = "APLICACION" if subcat == "aplicacion" else "UNION"

        filename = GSHEET_APP_TOKEN  # siempre token, no archivo

        return render_template(
            "line_start.html",
            line_key=line_key,
            line_title=ui["title"],
            line_desc=ui["desc"],
            stage="select_sheet",
            filename=filename,
            machine=machine,
            subcat=subcat,
            machine_id=machine_id,
        )

    # -------------------------
    # ✅ CONTINUIDAD (SIEMPRE Google Sheets)
    # -------------------------
    if line_key == "continuidad":
        if machine not in ("BUSES", "MOTOS"):
            flash("Debes seleccionar Línea de motos o Línea de buses.", "error")
            return redirect(url_for("line_start", line_key=line_key))

        filename = GSHEET_CONT_BUSES_TOKEN if machine == "BUSES" else GSHEET_CONT_MOTOS_TOKEN

        return render_template(
            "line_start.html",
            line_key=line_key,
            line_title=ui["title"],
            line_desc=ui["desc"],
            stage="select_sheet",
            filename=filename,
            machine=machine,
            subcat=subcat,
            machine_id=machine_id,
        )

    # -------------------------
    # ✅ fallback
    # -------------------------
    flash("Línea no soportada.", "error")
    return redirect(url_for("home"))


def _week_label_with_range(y: int, w: int) -> str:
    try:
        start = pd.to_datetime(f"{int(y)}-W{int(w):02d}-1", format="%G-W%V-%u", errors="coerce")
        if pd.isna(start):
            return f"{int(y)} - Semana {int(w):02d}"

        end = start + pd.Timedelta(days=6)
        return f"{int(y)} - Semana {int(w):02d} [{start.strftime('%d/%m')} - {end.strftime('%d/%m')}]"
    except Exception:
        return f"{int(y)} - Semana {int(w):02d}"


@app.get("/period_options")
def period_options():
    analysis_key = (request.args.get("analysis_key") or "").strip().lower()
    filename = (request.args.get("filename") or "").strip()
    machine = (request.args.get("machine") or "").strip().upper()
    period = (request.args.get("period") or "day").strip().lower()
    machine_id = (request.args.get("machine_id") or "").strip()  # ✅ NUEVO

    # ✅ Ciclo/Aplicadores: NO usa day/week/month. Solo TOTAL.
    if analysis_key == "ciclo_aplicadores":
        return jsonify({"options": [{"value": "TOTAL", "label": "TOTAL"}]})

    # ✅ CONTINUIDAD (GSHEET)
    if analysis_key == "continuidad":
        try:
            ttl = int(getattr(config, "GSHEET_TTL_S", 10) or 10)

            m = (machine or "").strip().upper()
            if m == "BUSES":
                sheet_id = getattr(config, "CONT_BUSES_SHEET_ID", "") or ""
                gid_hora = int(getattr(config, "CONT_BUSES_GID_RESUMEN_HORA", 0) or 0)
            elif m == "MOTOS":
                sheet_id = getattr(config, "CONT_MOTOS_SHEET_ID", "") or ""
                gid_hora = int(getattr(config, "CONT_MOTOS_GID_RESUMEN_HORA", 0) or 0)
            else:
                # si no viene máquina bien, no inventamos
                return jsonify({"options": []})

            if not sheet_id or gid_hora is None:
                return jsonify({"options": []})

            df = load_gsheet_tab_csv_smart(sheet_id, gid_hora, ttl_s=ttl)

            c_fecha = _find_col(df, "fecha")
            c_hora = _find_col(df, "hora")
            if not c_fecha or not c_hora:
                return jsonify({"options": []})

            dt = _build_dt_from_fecha_hora(df, c_fecha, c_hora)
            dt = pd.to_datetime(dt, errors="coerce").dropna()
            if dt.empty:
                return jsonify({"options": []})

            if period == "day":
                days = sorted(dt.dt.normalize().unique(), reverse=True)
                options = [{"value": d.strftime("%Y-%m-%d"), "label": d.strftime("%d/%m/%Y")} for d in days]
                return jsonify({"options": options})

            if period == "week":
                iso = dt.dt.isocalendar()
                uniq = sorted(set(zip(iso["year"].astype(int), iso["week"].astype(int))), reverse=True)
                options = [
                    {
                        "value": f"{y}-W{w:02d}",
                        "label": _week_label_with_range(int(y), int(w))
                    }
                    for (y, w) in uniq
                ]
                return jsonify({"options": options})

            if period == "month":
                uniq = sorted(dt.dt.strftime("%Y-%m").unique(), reverse=True)
                options = [{"value": x, "label": x} for x in uniq]
                return jsonify({"options": options})

            return jsonify({"options": []})

        except Exception:
            app.logger.exception("period_options CONTINUIDAD failed")
            return jsonify({"options": []})

    # ✅ APLICACION / UNION: depende del machine_id (Máquina 1/2) para mostrar fechas activas correctas
    if machine in ("APLICACION", "UNION"):
        try:
            path = GSHEET_APP_TOKEN if filename == GSHEET_APP_TOKEN else os.path.join(app.config["UPLOAD_FOLDER"],
                                                                                      filename)

            try:
                options = get_period_options_aplicacion(path, period, machine=machine, machine_id=machine_id)
            except TypeError:
                options = get_period_options_aplicacion(path, period)

            return jsonify({"options": options})
        except Exception:
            app.logger.exception("period_options APLICACION/UNION failed")
            return jsonify({"options": []})

    # ✅ CORTE: soporta Excel + Google Sheets
    is_hp_gsheet = (machine == "HP" and filename == GSHEET_HP_TOKEN)
    is_thb_gsheet = (machine == "THB" and filename == GSHEET_THB_TOKEN)

    # hoja base por máquina
    base_sheet = "Corte" if machine == "HP" else "Verificación de Medidas"

    try:
        if is_hp_gsheet:
            sheet_id = getattr(config, "HP_SHEET_ID", "") or ""
            gid_corte = int(getattr(config, "HP_GID_CORTE", 0) or 0)
            ttl = int(getattr(config, "GSHEET_TTL_S", 10) or 10)
            if not sheet_id or not gid_corte:
                return jsonify({"options": []})
            df = load_gsheet_tab_csv_smart(sheet_id, gid_corte, ttl_s=ttl)

        elif is_thb_gsheet:
            sheet_id = getattr(config, "THB_SHEET_ID", "") or ""
            gid_verif = int(getattr(config, "THB_GID_VERIF", 0) or 0)
            ttl = int(getattr(config, "GSHEET_TTL_S", 10) or 10)
            if not sheet_id or not gid_verif:
                return jsonify({"options": []})
            df = load_gsheet_tab_csv_smart(sheet_id, gid_verif, ttl_s=ttl)

        else:
            path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
            if not filename or not os.path.exists(path):
                return jsonify({"options": []})

            try:
                df = _read_excel_smart(path, base_sheet)
            except Exception:
                # fallback por acento
                df = _read_excel_smart(path, "Verificacion de Medidas")

    except Exception:
        app.logger.exception("period_options CORTE failed")
        return jsonify({"options": []})

    c_fecha = _find_col(df, "fecha")
    c_hora = _find_col(df, "hora")
    if not c_fecha or not c_hora:
        return jsonify({"options": []})

    dt = _build_dt_from_fecha_hora(df, c_fecha, c_hora)
    dt = pd.to_datetime(dt, errors="coerce").dropna()
    if dt.empty:
        return jsonify({"options": []})

    if period == "day":
        days = sorted(dt.dt.normalize().unique(), reverse=True)
        options = [{"value": d.strftime("%Y-%m-%d"), "label": d.strftime("%d/%m/%Y")} for d in days]
        return jsonify({"options": options})

    if period == "week":
        iso = dt.dt.isocalendar()
        uniq = sorted(set(zip(iso["year"].astype(int), iso["week"].astype(int))), reverse=True)
        options = [
            {
                "value": f"{y}-W{w:02d}",
                "label": _week_label_with_range(int(y), int(w))
            }
            for (y, w) in uniq
        ]
        return jsonify({"options": options})

    if period == "month":
        uniq = sorted(dt.dt.strftime("%Y-%m").unique(), reverse=True)
        options = [{"value": x, "label": x} for x in uniq]
        return jsonify({"options": options})

    return jsonify({"options": []})


@app.get("/thb_tarjetas_hour")
def thb_tarjetas_hour():
    import os
    import re
    import pandas as pd
    from flask import request, jsonify

    filename = (request.args.get("filename", "") or "").strip()
    period_value = (request.args.get("period_value", "") or "").strip()
    hour_start = (request.args.get("hour_start", "") or "").strip()
    hour_end = (request.args.get("hour_end", "") or "").strip()
    operator = (request.args.get("operator", "") or "").strip()
    limit = int((request.args.get("limit", "250") or "250").strip() or 250)

    def _nocache_json(payload, status=200):
        resp = jsonify(payload)
        resp.status_code = status
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
        return resp

    if not period_value or not hour_start or not hour_end:
        return _nocache_json(
            {"ok": False, "error": "Faltan parámetros (period_value/hour_start/hour_end)."},
            400
        )

    def _parse_day(pv: str):
        pv2 = (pv or "").strip().replace("/", "-")
        if re.match(r"^\d{4}-\d{2}-\d{2}$", pv2):
            return pd.to_datetime(pv2, errors="coerce", format="%Y-%m-%d")
        if re.match(r"^\d{2}-\d{2}-\d{4}$", pv2):
            return pd.to_datetime(pv2, errors="coerce", format="%d-%m-%Y")
        return pd.to_datetime(pv2, errors="coerce", dayfirst=True)

    def _parse_hhmm(x: str):
        m = re.search(r"(\d{1,2})\s*:\s*(\d{1,2})", str(x or ""))
        if not m:
            return None
        hh = int(m.group(1));
        mm = int(m.group(2))
        if hh < 0 or hh > 23 or mm < 0 or mm > 59:
            return None
        return hh, mm

    d0 = _parse_day(period_value)
    if pd.isna(d0):
        return _nocache_json({"ok": False, "error": "period_value inválido."}, 400)

    hs = _parse_hhmm(hour_start)
    he = _parse_hhmm(hour_end)
    if not hs or not he:
        return _nocache_json({"ok": False, "error": "hour_start/hour_end inválidos."}, 400)

    start_dt = pd.Timestamp(d0).normalize() + pd.Timedelta(hours=hs[0], minutes=hs[1])
    end_dt = pd.Timestamp(d0).normalize() + pd.Timedelta(hours=he[0], minutes=he[1])
    if end_dt <= start_dt:
        end_dt += pd.Timedelta(days=1)

    # =========================================================
    # Cargar "Verificación de Medidas" (cache → gsheet → excel)
    # =========================================================
    df_verif = None
    try:
        cache = THB_CACHE.get(filename, None)
        if isinstance(cache, dict) and isinstance(cache.get("df"), pd.DataFrame):
            df_verif = cache["df"].copy()
        else:
            if filename == GSHEET_THB_TOKEN:
                sheet_id = getattr(config, "THB_SHEET_ID", "") or ""
                gid_verif = int(getattr(config, "THB_GID_VERIF", 0) or 0)
                ttl = int(getattr(config, "GSHEET_TTL_S", 10) or 10)

                if not sheet_id or not gid_verif:
                    return _nocache_json(
                        {"ok": False, "error": "Falta THB_SHEET_ID/THB_GID_VERIF en config.py."},
                        400
                    )

                df_verif = load_gsheet_tab_csv_smart(sheet_id, gid_verif, ttl_s=ttl).copy()
            else:
                path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                if not filename or not os.path.exists(path):
                    return _nocache_json({"ok": False, "error": "Archivo no existe."}, 400)

                try:
                    df_verif = _read_excel_smart(path, "Verificación de Medidas").copy()
                except Exception:
                    df_verif = _read_excel_smart(path, "Verificacion de Medidas").copy()

    except Exception as e:
        return _nocache_json({"ok": False, "error": f"No pude cargar Verificación de Medidas: {e}"}, 500)

    if df_verif is None or df_verif.empty:
        return _nocache_json({"ok": True, "total": 0, "items": [], "truncated": False}, 200)

    # =========================================================
    # Columnas requeridas
    # =========================================================
    c_fecha = _find_col(df_verif, "fecha")
    c_hora = _find_col(df_verif, "hora")
    c_consec = _find_col(df_verif, "consecutivo")
    c_codigo = (
            _find_col(df_verif, "código") or _find_col(df_verif, "codigo") or
            _find_col(df_verif, "arnés") or _find_col(df_verif, "arnes")
    )
    c_nombre = _find_col(df_verif, "nombre")

    if not c_fecha or not c_hora or not c_consec or not c_codigo:
        return _nocache_json(
            {"ok": False, "error": "No encuentro columnas (Fecha/Hora/Consecutivo/Código del arnés)."},
            400
        )

    dt = _build_dt_from_fecha_hora(df_verif, c_fecha, c_hora)
    dt = pd.to_datetime(dt, errors="coerce")

    mask = dt.notna() & (dt >= start_dt) & (dt < end_dt)

    # opcional: filtrar por operaria si viene diferente a General
    if c_nombre and operator and operator.strip() and operator.strip().lower() != "general":
        mask = mask & (df_verif[c_nombre].astype(str).str.strip() == operator.strip())

    df = df_verif.loc[mask, [c_consec, c_codigo]].copy()
    if df.empty:
        return _nocache_json(
            {"ok": True, "total": 0, "items": [], "truncated": False,
             "hour_start": hour_start, "hour_end": hour_end, "period_value": period_value},
            200
        )

    def _clean(v):
        if pd.isna(v):
            return ""
        s = str(v).strip()
        if s.lower() in ("nan", "none"):
            return ""
        return s

    items = []
    for _, r in df.iterrows():
        cons = _clean(r[c_consec])
        cod = _clean(r[c_codigo])

        # normalizar consecutivo tipo "2.0" -> "2"
        try:
            if cons and re.match(r"^\d+\.0$", cons):
                cons = str(int(float(cons)))
        except Exception:
            pass

        if cons or cod:
            items.append({"consecutivo": cons, "codigo_arnes": cod})

    total = len(items)
    truncated = total > limit
    items = items[:limit]

    return _nocache_json(
        {
            "ok": True,
            "period_value": period_value,
            "hour_start": hour_start,
            "hour_end": hour_end,
            "total": total,
            "limit": limit,
            "truncated": truncated,
            "items": items,
        },
        200
    )


@app.get("/thb_filter_options")
def thb_filter_options():
    import os
    import pandas as pd
    from flask import request, jsonify

    filename = request.args.get("filename", "")
    machine = (request.args.get("machine", "THB") or "THB").strip().upper()
    period = (request.args.get("period", "day") or "day").strip().lower()
    pv = (request.args.get("period_value", "") or "").strip()

    # ✅ tokens Google Sheets
    is_hp_gsheet = (machine == "HP" and filename == GSHEET_HP_TOKEN)
    is_thb_gsheet = (machine == "THB" and filename == GSHEET_THB_TOKEN)

    # ✅ path solo para Excel local
    if not (is_hp_gsheet or is_thb_gsheet):
        path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    else:
        path = ""  # no se usa en modo gsheet

    def _sec_to_hhmm(s: int) -> str:
        s = int(max(0, s))
        hh = s // 3600
        mm = (s % 3600) // 60
        return f"{hh:02d}:{mm:02d}"

    def _filter_dt_by_period(dt: pd.Series, period: str, pv: str) -> pd.Series:
        m = dt.notna()
        if not pv:
            return m

        if period == "day":
            # pv puede venir "YYYY-MM-DD" o "DD/MM/YYYY" (o "DD-MM-YYYY")
            pv2 = (pv or "").strip().replace("/", "-")

            # Detectar formato para evitar warning dayfirst con YYYY-MM-DD
            if re.match(r"^\d{4}-\d{2}-\d{2}$", pv2):
                d0 = pd.to_datetime(pv2, errors="coerce", format="%Y-%m-%d")
            elif re.match(r"^\d{2}-\d{2}-\d{4}$", pv2):
                d0 = pd.to_datetime(pv2, errors="coerce", format="%d-%m-%Y")
            else:
                d0 = pd.to_datetime(pv2, errors="coerce", dayfirst=True)

            if not pd.isna(d0):
                m = m & (dt.dt.normalize() == pd.Timestamp(d0).normalize())


        elif period == "month":
            # pv esperado "YYYY-MM"
            m = m & (dt.dt.strftime("%Y-%m") == pv)

        elif period == "week":
            # pv esperado "YYYY-Www"
            try:
                pv2 = pv.replace("w", "W")
                y = int(pv2.split("-W")[0])
                w = int(pv2.split("-W")[1])
                iso = dt.dt.isocalendar()
                m = m & (iso["year"].astype(int) == y) & (iso["week"].astype(int) == w)
            except Exception:
                pass

        return m

    # =========================================================
    # ✅ 1) Cargar df_verif según máquina y fuente (GSheet/Excel)
    # =========================================================
    if is_hp_gsheet:
        # HP: verif desde HP sheet
        try:
            sheet_id = getattr(config, "HP_SHEET_ID", "") or ""
            gid_verif = int(getattr(config, "HP_GID_VERIF", 0) or 0)
            ttl = int(getattr(config, "GSHEET_TTL_S", 10) or 10)
            if not sheet_id or not gid_verif:
                return jsonify({"operators": [], "min_time": "00:00", "max_time": "23:59"})
            df_verif = load_gsheet_tab_csv_smart(sheet_id, gid_verif, ttl_s=ttl)
        except Exception:
            return jsonify({"operators": [], "min_time": "00:00", "max_time": "23:59"})

    elif is_thb_gsheet:
        # THB: verif desde THB sheet
        try:
            sheet_id = getattr(config, "THB_SHEET_ID", "") or ""
            gid_verif = int(getattr(config, "THB_GID_VERIF", 0) or 0)
            ttl = int(getattr(config, "GSHEET_TTL_S", 10) or 10)
            if not sheet_id or not gid_verif:
                return jsonify({"operators": [], "min_time": "00:00", "max_time": "23:59"})
            df_verif = load_gsheet_tab_csv_smart(sheet_id, gid_verif, ttl_s=ttl)
        except Exception:
            return jsonify({"operators": [], "min_time": "00:00", "max_time": "23:59"})

    else:
        # Excel local (THB o HP)
        cache_thb = THB_CACHE.get(filename)
        if cache_thb and isinstance(cache_thb.get("df"), pd.DataFrame):
            df_verif = cache_thb["df"]
        else:
            sheet_verif = "Verificación de Medidas"
            try:
                try:
                    df_verif = _read_excel_smart(path, sheet_verif)
                except Exception:
                    df_verif = _read_excel_smart(path, "Verificacion de Medidas")
            except Exception:
                return jsonify({"operators": [], "min_time": "00:00", "max_time": "23:59"})

    # =========================================================
    # ✅ 2) Columnas base + filtro periodo + operadores
    # =========================================================
    c_fecha_v = _find_col(df_verif, "fecha")
    c_hora_v = _find_col(df_verif, "hora")
    c_nom_v = _find_col(df_verif, "nombre")

    if not c_fecha_v or not c_hora_v:
        return jsonify({"operators": [], "min_time": "00:00", "max_time": "23:59"})

    dt_verif = _build_dt_from_fecha_hora(df_verif, c_fecha_v, c_hora_v)
    m_verif = _filter_dt_by_period(dt_verif, period, pv)

    dfv2 = df_verif.loc[m_verif].copy()

    ops = []
    if c_nom_v and c_nom_v in dfv2.columns:
        ops = sorted(set(dfv2[c_nom_v].astype(str).str.strip().replace({"nan": "", "None": ""})))
        ops = [o for o in ops if o and o.lower() != "general"]

    # ✅ semana/mes => NO mostrar horas
    if period in ("week", "month"):
        return jsonify({"operators": ops, "min_time": "", "max_time": ""})

    # =========================================================
    # ✅ 3) THB: min/max hora desde verif
    # =========================================================
    if machine == "THB":
        dt2 = dt_verif.loc[m_verif].dropna()
        if dt2.empty:
            return jsonify({"operators": ops, "min_time": "00:00", "max_time": "23:59"})

        tod = (dt2 - dt2.dt.normalize()).dt.total_seconds()
        min_s = int(tod.min())
        max_s = int(tod.max())
        # ✅ Entrada del turno THB: 07:15. Si la primera actividad es anterior,
        #    se ancla a 07:15 (no se paga antes de la entrada).
        SHIFT_START_S = 7 * 3600 + 15 * 60
        if min_s < SHIFT_START_S:
            min_s = SHIFT_START_S
        return jsonify({"operators": ops, "min_time": _sec_to_hhmm(min_s), "max_time": _sec_to_hhmm(max_s)})

    # =========================================================
    # ✅ 4) HP: min/max hora desde Corte (+ Paradas si existe)
    # =========================================================
    if machine == "HP":
        dtc2 = pd.Series([], dtype="datetime64[ns]")  # ✅ evita UnboundLocalError

        # ---- leer corte / paradas (Excel o GSheet) ----
        try:
            if is_hp_gsheet:
                sheet_id = getattr(config, "HP_SHEET_ID", "") or ""
                gid_corte = int(getattr(config, "HP_GID_CORTE", 0) or 0)
                gid_par = int(getattr(config, "HP_GID_PARADAS", 0) or 0)
                ttl = int(getattr(config, "GSHEET_TTL_S", 10) or 10)

                if not sheet_id or not gid_corte:
                    return jsonify({"operators": ops, "min_time": "00:00", "max_time": "23:59"})

                df_corte = load_gsheet_tab_csv_smart(sheet_id, gid_corte, ttl_s=ttl)

                df_par = None
                if gid_par:
                    try:
                        df_par = load_gsheet_tab_csv_smart(sheet_id, gid_par, ttl_s=ttl)
                    except Exception:
                        df_par = None

            else:
                # Excel local
                path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                try:
                    df_corte = _read_excel_smart(path, "Corte")
                except Exception:
                    return jsonify({"operators": ops, "min_time": "00:00", "max_time": "23:59"})

                try:
                    df_par = _read_excel_smart(path, "Paradas")
                except Exception:
                    df_par = None

        except Exception:
            return jsonify({"operators": ops, "min_time": "00:00", "max_time": "23:59"})

        # ---- dt corte ----
        c_fecha_c = _find_col(df_corte, "fecha")
        c_hora_c = _find_col(df_corte, "hora")
        if not c_fecha_c or not c_hora_c:
            return jsonify({"operators": ops, "min_time": "00:00", "max_time": "23:59"})

        dt_corte = _build_dt_from_fecha_hora(df_corte, c_fecha_c, c_hora_c)
        m_corte = _filter_dt_by_period(dt_corte, period, pv)
        dtc2 = dt_corte.loc[m_corte].dropna()

        # ---- sumar paradas (para min/max) ----
        if df_par is not None and isinstance(df_par, pd.DataFrame) and not df_par.empty:
            c_fecha_p = _find_col(df_par, "fecha")
            c_hora_p = _find_col(df_par, "hora")
            if c_fecha_p and c_hora_p:
                dt_par = _build_dt_from_fecha_hora(df_par, c_fecha_p, c_hora_p)
                m_par = _filter_dt_by_period(dt_par, period, pv)
                dtp2 = dt_par.loc[m_par].dropna()
                if not dtp2.empty:
                    if dtc2.empty:
                        dtc2 = dtp2
                    else:
                        dtc2 = pd.concat([dtc2, dtp2], ignore_index=True)

        if dtc2.empty:
            return jsonify({"operators": ops, "min_time": "00:00", "max_time": "23:59"})

        tod = (dtc2 - dtc2.dt.normalize()).dt.total_seconds()
        min_s = int(tod.min())
        max_s = int(tod.max())
        # ✅ Entrada del turno HP: 07:15. Si la primera actividad es anterior,
        #    se ancla a 07:15 (la capacidad no cuenta antes de la entrada).
        SHIFT_START_S = 7 * 3600 + 15 * 60
        # HP siempre arranca a las 07:15. El selector no debe usar 07:42, 08:10, etc.
        # como inicio automático, porque recortaría paradas que terminaron después
        # pero empezaron desde la entrada del turno.
        min_s = SHIFT_START_S
        return jsonify({"operators": ops, "min_time": _sec_to_hhmm(min_s), "max_time": _sec_to_hhmm(max_s)})

    # fallback
    return jsonify({"operators": ops, "min_time": "00:00", "max_time": "23:59"})


@app.route("/app_filter_options", methods=["GET"])
def app_filter_options():
    filename = request.args.get("filename", "")
    machine = (request.args.get("machine", "") or "").strip().upper()
    subcat = (request.args.get("subcat", "") or "").strip().lower()
    period = request.args.get("period", "day")
    period_value = request.args.get("period_value", "")
    machine_id = (request.args.get("machine_id") or "").strip()  # ✅ NUEVO

    if not filename or not period_value:
        return jsonify({"operators": [], "min_time": "", "max_time": ""})

    if machine not in ("APLICACION", "UNION"):
        if subcat in ("aplicacion", "union"):
            machine = "APLICACION" if subcat == "aplicacion" else "UNION"
        else:
            return jsonify({"operators": [], "min_time": "", "max_time": ""})

    if filename == GSHEET_APP_TOKEN:
        path = GSHEET_APP_TOKEN
    else:
        path = safe_upload_path(app.config["UPLOAD_FOLDER"], filename)
        if not os.path.exists(path):
            return jsonify({"operators": [], "min_time": "", "max_time": ""})

    try:
        from analyses.aplicacion_excel import get_app_filter_options_aplicacion

        # ✅ Compat: si tu función aún no recibe machine_id, no rompe
        try:
            data = get_app_filter_options_aplicacion(
                path=path,
                period=period,
                period_value=period_value,
                machine=machine,
                subcat=subcat,
                machine_id=machine_id,
            )
        except TypeError:
            data = get_app_filter_options_aplicacion(
                path=path,
                period=period,
                period_value=period_value,
                machine=machine,
                subcat=subcat,
            )

        # ✅ MOD PRINCIPAL: semana/mes => NO mostrar horas
        p = (period or "").strip().lower()
        if p in ("week", "month"):
            min_time = ""
            max_time = ""
        else:
            min_time = data.get("min_time", "00:00")
            max_time = data.get("max_time", "23:59")

        return jsonify({
            "operators": data.get("operators", []),
            "min_time": min_time,
            "max_time": max_time,
        })
    except Exception:
        p = (period or "").strip().lower()
        if p in ("week", "month"):
            return jsonify({"operators": [], "min_time": "", "max_time": ""})
        return jsonify({"operators": [], "min_time": "00:00", "max_time": "23:59"})


from analyses.corte_thb import (
    analyze_thb,
    _resolve_thb_cols,
    _build_dt as _thb_build_dt,
    _num_es as _thb_num_es,
    _normalize_period_value as _thb_normalize_period_value,
    _period_bounds as _thb_period_bounds,
)
from analyses.corte_hp import (
    analyze_hp,
    _resolve_hp_corte_cols,
    _build_dt as _hp_build_dt,
    _range_from_period as _hp_range_from_period,
)

import io
import urllib.parse
import urllib.request
import pandas as pd


def _load_gsheet_tab_by_name(sheet_id: str, sheet_name: str) -> pd.DataFrame:
    # Descarga una pestaña por nombre como CSV (no necesita gid)
    url = (
            "https://docs.google.com/spreadsheets/d/"
            + str(sheet_id)
            + "/gviz/tq?tqx=out:csv&sheet="
            + urllib.parse.quote(sheet_name)
    )
    b = urllib.request.urlopen(url).read()
    return pd.read_csv(io.BytesIO(b), header=None)  # header=None porque tú quieres columna D fija


# =========================
# CICLO/APLICADORES — UNION (4 libros) — Col D desde fila 3
# =========================
_CICLO_UNION_CACHE = {"ts": 0.0, "df": None}

# cache simple (evita pegarle al sheet en cada request)
_UNION_RAW_CACHE: dict[tuple[str, int], tuple[float, pd.DataFrame]] = {}


def _load_union_pulsos_4books_core(ttl_s: int | None = None) -> pd.DataFrame:
    """
    CICLO/APLICADORES - Uniones.

    Compatibilidad:
    - Libros viejos: hoja/resumen con columnas Terminal y Pulsos.
    - Libros nuevos UNION 1/2: UNION*_Resumen_Hora, donde el GID de resumen
      está en GSHEET_GID_APLICACION3 / GSHEET_GID_APLICACION4 y las paradas
      quedan en GSHEET_GID_UNION3 / GSHEET_GID_UNION4.

    Nunca usa una hoja de paradas si no encuentra explícitamente encabezados
    'Terminal' y 'Pulsos'. Así se evita sumar columnas equivocadas.
    """
    import pandas as pd
    import time
    import re

    ttl = int(getattr(config, "GSHEET_TTL_S", 10) or 10) if ttl_s is None else int(ttl_s)
    now = time.time()

    hit = _CICLO_UNION_CACHE.get("df")
    if hit is not None and (now - float(_CICLO_UNION_CACHE.get("ts", 0.0))) < ttl:
        return hit.copy()

    def _gid_list(*values):
        out = []
        for v in values:
            try:
                gi = int(v or 0)
            except Exception:
                gi = 0
            if gi > 0 and gi not in out:
                out.append(gi)
        return out

    def _download_gid(sheet_id: str, gid: int) -> pd.DataFrame:
        sid = str(sheet_id or "").strip()
        gid_i = int(gid or 0)
        if not sid or not gid_i:
            return pd.DataFrame()

        urls = [
            f"https://docs.google.com/spreadsheets/d/{sid}/export?format=csv&gid={gid_i}",
            f"https://docs.google.com/spreadsheets/d/{sid}/gviz/tq?tqx=out:csv&gid={gid_i}",
        ]
        last = None
        for url in urls:
            try:
                r = requests.get(
                    url,
                    headers={"User-Agent": "Mozilla/5.0 PlataformaMaquinas/1.0"},
                    timeout=25,
                )
                r.raise_for_status()
                raw = pd.read_csv(io.StringIO(r.text), header=None)
                return raw.dropna(how="all").reset_index(drop=True)
            except Exception as e:
                last = e
                continue
        raise RuntimeError(f"No se pudo leer UNION CSV gid={gid_i}: {last}")

    def _find_header_row(raw: pd.DataFrame, max_scan: int = 15) -> int | None:
        n = min(max_scan, len(raw))
        for i in range(n):
            row = raw.iloc[i].astype(str).str.strip().str.lower().tolist()
            if any("puls" in c for c in row) and any("terminal" in c for c in row):
                return i
        return None

    def _col_idx_by_name(header_row: list[str], key: str) -> int | None:
        k = (key or "").strip().lower()
        for idx, cell in enumerate(header_row):
            s = str(cell).strip().lower()
            if k in s:
                return idx
        return None

    def _parse_pulses(x) -> int:
        s = str(x or "").strip().replace("\u00a0", "").replace(" ", "")
        if not s or s.upper() in ("#VALUE!", "#REF!", "#DIV/0!", "#N/A", "N/A", "NA"):
            return 0
        if re.match(r"^\d{1,3}(\.\d{3})+(,\d+)?$", s):
            s = s.replace(".", "").replace(",", ".")
        elif re.match(r"^\d{1,3}(,\d{3})+(\.\d+)?$", s):
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
        try:
            return int(round(float(s)))
        except Exception:
            return 0

    books = [
        # U1/U2 conservan compatibilidad con el tab viejo de Unión.
        ("U1", getattr(config, "GSHEET_ID", None),
         _gid_list(getattr(config, "GSHEET_GID_UNION", None), getattr(config, "GSHEET_GID_APLICACION", None))),
        ("U2", getattr(config, "GSHEET_ID2", None),
         _gid_list(getattr(config, "GSHEET_GID_UNION2", None), getattr(config, "GSHEET_GID_APLICACION2", None))),
        # U3/U4 son las máquinas de Unión nuevas: primero Resumen_Hora, luego paradas como respaldo.
        ("U3", getattr(config, "GSHEET_ID3", None),
         _gid_list(getattr(config, "GSHEET_GID_APLICACION3", None), getattr(config, "GSHEET_GID_UNION3", None))),
        ("U4", getattr(config, "GSHEET_ID4", None),
         _gid_list(getattr(config, "GSHEET_GID_APLICACION4", None), getattr(config, "GSHEET_GID_UNION4", None))),
    ]

    frames: list[pd.DataFrame] = []

    for tag, sid, gids in books:
        if not sid or not gids:
            continue

        selected = None
        for gid in gids:
            try:
                raw = _download_gid(str(sid), int(gid))
            except Exception:
                continue
            if raw.empty:
                continue

            h = _find_header_row(raw)
            if h is None:
                continue

            header = raw.iloc[h].astype(str).str.strip().tolist()
            col_pulsos = _col_idx_by_name(header, "puls")
            col_term = _col_idx_by_name(header, "terminal")
            if col_pulsos is None or col_term is None:
                continue

            selected = (raw, h, col_pulsos, col_term, gid)
            break

        if selected is None:
            continue

        raw, h, col_pulsos, col_term, gid_used = selected
        df0 = raw.iloc[h + 1:].copy().reset_index(drop=True)
        if df0.shape[1] <= max(col_pulsos, col_term):
            continue

        pulsos = df0.iloc[:, col_pulsos].apply(_parse_pulses).astype("int64")
        terminal = df0.iloc[:, col_term].astype(str).str.strip()
        terminal = terminal.replace({"nan": "", "None": "", "#VALUE!": "", "#N/A": "", "#REF!": ""})
        terminal = terminal.str.replace(r"\.0$", "", regex=True).str.strip()

        tmp = pd.DataFrame({"terminal": terminal, "pulsos": pulsos})
        tmp["src"] = tag
        tmp["gid"] = int(gid_used)
        tmp = tmp[tmp["terminal"].ne("")]
        tmp = tmp[~tmp["terminal"].str.lower().isin(["total", "totales"])]
        tmp = tmp[tmp["terminal"].str.match(r"^\d+$", na=False)]
        tmp = tmp[tmp["pulsos"] > 0]
        if not tmp.empty:
            frames.append(tmp)

    out = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=["terminal", "pulsos", "src", "gid"])
    _CICLO_UNION_CACHE["ts"] = now
    _CICLO_UNION_CACHE["df"] = out.copy()
    return out.copy()


def _load_union_pulsos_4books() -> pd.DataFrame:
    return _load_union_pulsos_4books_core()


# =========================
# EXPORT THB -> Excel por día / semana / mes
# =========================
def _load_thb_data_for_export(filename: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    cache_thb = THB_CACHE.get(filename)

    if filename == GSHEET_THB_TOKEN:
        sheet_id = getattr(config, "THB_SHEET_ID", "") or ""
        gid_verif = int(getattr(config, "THB_GID_VERIF", 0) or 0)
        gid_par = int(getattr(config, "THB_GID_PARADAS", 0) or 0)
        ttl = int(getattr(config, "GSHEET_TTL_S", 10) or 10)

        if not sheet_id or not gid_verif or not gid_par:
            raise RuntimeError("Falta config THB en config.py (THB_SHEET_ID/THB_GID_*).")

        df_verif = load_gsheet_tab_csv_smart(sheet_id, gid_verif, ttl_s=ttl).copy()
        df_paradas = load_gsheet_tab_csv_smart(sheet_id, gid_par, ttl_s=ttl).copy()
        df_paradas = _normalize_gsheet_paradas_duration_to_seconds(df_paradas)
        df_paradas = _hp_gsheet_fix_paradas_duration(df_paradas)

        try:
            if filename in THB_CACHE and isinstance(THB_CACHE[filename], dict):
                THB_CACHE[filename]["df"] = df_verif.copy()
                THB_CACHE[filename]["df_paradas"] = df_paradas.copy()
            else:
                THB_CACHE[filename] = {"df": df_verif.copy(), "df_paradas": df_paradas.copy()}
        except Exception:
            pass

        return df_verif, df_paradas

    path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    if not os.path.exists(path):
        raise FileNotFoundError(f"No existe el archivo: {filename}")

    if cache_thb and isinstance(cache_thb.get("df"), pd.DataFrame):
        df_verif = cache_thb["df"].copy()
        df_paradas = cache_thb.get("df_paradas", pd.DataFrame()).copy()
        return df_verif, df_paradas

    try:
        df_verif = _read_excel_smart(path, "Verificación de Medidas").copy()
    except Exception:
        df_verif = _read_excel_smart(path, "Verificacion de Medidas").copy()

    df_paradas = _read_excel_smart(path, "Paradas").copy()

    try:
        if filename in THB_CACHE and isinstance(THB_CACHE[filename], dict):
            THB_CACHE[filename]["df"] = df_verif.copy()
            THB_CACHE[filename]["df_paradas"] = df_paradas.copy()
        else:
            THB_CACHE[filename] = {"df": df_verif.copy(), "df_paradas": df_paradas.copy()}
    except Exception:
        pass

    return df_verif, df_paradas


# =========================
# EXPORT HP -> Excel por día / semana / mes
# =========================
def _load_hp_data_for_export(filename: str) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Carga SOLO las hojas persistentes de HP para exportar:
      - Corte
      - Verificación de Medidas / Verificacion de Medidas
      - Paradas
      - Duración / Duracion (opcional en Excel local)

    Importante: NO usa la hoja temporal "HP".
    """
    if filename == GSHEET_HP_TOKEN:
        sheet_id = getattr(config, "HP_SHEET_ID", "") or ""
        gid_corte = int(getattr(config, "HP_GID_CORTE", 0) or 0)
        gid_verif = int(getattr(config, "HP_GID_VERIF", 0) or 0)
        gid_par = int(getattr(config, "HP_GID_PARADAS", 0) or 0)
        ttl = int(getattr(config, "GSHEET_TTL_S", 10) or 10)

        if not sheet_id or not gid_corte or not gid_verif or not gid_par:
            raise RuntimeError("Falta config HP en config.py (HP_SHEET_ID/HP_GID_*).")

        df_corte = load_gsheet_tab_csv_smart(sheet_id, gid_corte, ttl_s=ttl).copy()
        df_verif = load_gsheet_tab_csv_smart(sheet_id, gid_verif, ttl_s=ttl).copy()
        df_paradas = load_gsheet_tab_csv_smart(sheet_id, gid_par, ttl_s=ttl).copy()
        df_paradas = _normalize_gsheet_paradas_duration_to_seconds(df_paradas)
        df_paradas = _hp_gsheet_fix_paradas_duration(df_paradas)
        df_duracion = pd.DataFrame()
        return df_corte, df_verif, df_paradas, df_duracion

    path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    if not os.path.exists(path):
        raise FileNotFoundError(f"No existe el archivo: {filename}")

    df_corte = _read_excel_smart(path, "Corte").copy()

    try:
        df_verif = _read_excel_smart(path, "Verificación de Medidas").copy()
    except Exception:
        df_verif = _read_excel_smart(path, "Verificacion de Medidas").copy()

    df_paradas = _read_excel_smart(path, "Paradas").copy()

    try:
        df_duracion = pd.read_excel(path, sheet_name="Duración", header=None, engine="openpyxl")
    except Exception:
        try:
            df_duracion = pd.read_excel(path, sheet_name="Duracion", header=None, engine="openpyxl")
        except Exception:
            df_duracion = pd.DataFrame()

    return df_corte, df_verif, df_paradas, df_duracion


def _hp_days_for_period_export(df_corte: pd.DataFrame, period: str, period_value: str) -> list[pd.Timestamp]:
    cols = _resolve_hp_corte_cols(df_corte)
    if not cols.col_fecha:
        return []

    dt = _hp_build_dt(df_corte, cols.col_fecha, cols.col_hora)
    fallback = None
    if dt is not None and not pd.to_datetime(dt, errors="coerce").dropna().empty:
        fallback = pd.Timestamp(pd.to_datetime(dt, errors="coerce").dropna().min()).normalize()

    r0, r1 = _hp_range_from_period(period, period_value, fallback)
    if r0 is None or r1 is None:
        return []

    m = pd.to_datetime(dt, errors="coerce").notna() & (pd.to_datetime(dt, errors="coerce") >= r0) & (
                pd.to_datetime(dt, errors="coerce") <= r1)
    days = (
        pd.to_datetime(dt.loc[m], errors="coerce")
        .dropna()
        .dt.normalize()
        .drop_duplicates()
        .sort_values()
        .tolist()
    )
    return [pd.Timestamp(d) for d in days]


def _hp_export_rows_for_day(
        df_corte: pd.DataFrame,
        day: pd.Timestamp,
        buckets: list[dict],
        operator: str = "",
) -> list[dict]:
    cols = _resolve_hp_corte_cols(df_corte)
    if not cols.col_fecha:
        return []

    dt = _hp_build_dt(df_corte, cols.col_fecha, cols.col_hora)
    day0 = pd.Timestamp(day).normalize()
    day1 = day0 + pd.Timedelta(days=1)

    m = pd.to_datetime(dt, errors="coerce").notna() & (pd.to_datetime(dt, errors="coerce") >= day0) & (
                pd.to_datetime(dt, errors="coerce") < day1)
    sub = df_corte.loc[m].copy()
    dt_sub = pd.to_datetime(dt.loc[m], errors="coerce").copy()

    bucket_by_hour: dict[int, dict] = {}
    for b in (buckets or []):
        try:
            hh = int(str(b.get("hour") or "0"))
        except Exception:
            continue
        bucket_by_hour[hh] = b

    hours = sorted(bucket_by_hour.keys())
    if not hours and not sub.empty:
        hours = sorted(pd.to_datetime(dt_sub, errors="coerce").dropna().dt.hour.unique().tolist())

    def _num_series(df: pd.DataFrame, col: str | None) -> pd.Series:
        if not col or col not in df.columns:
            return pd.Series([], dtype="float64")
        return pd.to_numeric(
            df[col].astype(str)
            .str.replace("\u00a0", "", regex=False)
            .str.replace(" ", "", regex=False)
            .str.replace(r"^(\d{1,3})\.(\d{3})(,\d+)?$", lambda m: m.group(1) + m.group(2) + (m.group(3) or ""),
                         regex=True)
            .str.replace(",", ".", regex=False),
            errors="coerce"
        ).fillna(0.0).astype(float)

    rows: list[dict] = []
    op_fixed = (operator or "").strip()
    op_export = op_fixed if (op_fixed and op_fixed.lower() != "general") else "General"

    for hh in hours:
        h0 = day0 + pd.Timedelta(hours=int(hh))
        h1 = h0 + pd.Timedelta(hours=1)

        in_hour = dt_sub.notna() & (dt_sub >= h0) & (dt_sub < h1)
        hour_df = sub.loc[in_hour].copy()

        buenas = _num_series(hour_df, cols.col_buenas)
        total = _num_series(hour_df, cols.col_total)
        solic = _num_series(hour_df, cols.col_solic)

        if not buenas.empty:
            cut = int(round(float(buenas.sum())))
        elif not total.empty:
            cut = int(round(float(total.sum())))
        else:
            cut = 0

        if not total.empty:
            if not buenas.empty:
                planned_calc = int(round(float(total.where(total > 0, buenas).sum())))
            else:
                planned_calc = int(round(float(total.sum())))
        else:
            planned_calc = cut

        reject = int(round(float(solic.sum()))) if not solic.empty else max(0, int(cut) - int(planned_calc))

        referencias = []
        c_ref = cols.col_ref_cable
        if c_ref and c_ref in hour_df.columns and not hour_df.empty:
            vals = hour_df[c_ref].dropna().astype(str).str.strip()
            vals = vals[vals.ne("") & ~vals.str.lower().isin(["nan", "none"])]
            seen = set()
            for raw in vals.tolist():
                ref = str(raw).strip()
                key = ref.upper()
                if key in seen:
                    continue
                seen.add(key)
                referencias.append(ref)
                if len(referencias) >= 5:
                    break

        b = bucket_by_hour.get(int(hh), {})

        # Si el análisis del día ya trae buckets filtrados por operaria/ventana,
        # usar esos valores como fuente principal para que Excel y dashboard cierren igual.
        if b:
            try:
                cut = int(round(float(b.get("cut", cut) or 0)))
            except Exception:
                pass
            try:
                planned_calc = int(round(float(b.get("planned", b.get("planificado", planned_calc)) or 0)))
            except Exception:
                pass
            try:
                reject = int(round(float(b.get("nonconforming", b.get("reject", reject)) or 0)))
            except Exception:
                pass

        meal_sec = int(b.get("mealSec", 0) or 0)
        tx_sec = int(b.get("otherDeadSec", b.get("otherRecordedSec", 0)) or 0)

        # ✅ Regla HP: solo lunes, si existe parada 108, ese tiempo descuenta TP.
        # El bucket ya excluye 108 de TX cuando corresponde, pero protegemos el Excel
        # para que no castigue doble si viene de un resultado anterior.
        maint108_sec = int(b.get("maint108Sec", b.get("tpDiscount108Sec", 0)) or 0)
        try:
            is_monday = (pd.Timestamp(day0).weekday() == 0)
        except Exception:
            is_monday = False
        if not is_monday:
            maint108_sec = 0

        # Tiempo pagado FIJO y EXACTO por franja (no depende de la parada 105):
        #   - Hora 7  (07:00-08:00): entrada 07:15 -> se descuentan 15 min -> paga 0.75 h
        #   - Hora 8  (08:00-09:00): desayuno       -> se descuentan 15 min -> paga 0.75 h
        #   - Hora 13 (13:00-14:00): almuerzo       -> se descuentan 30 min -> paga 0.50 h
        #   - Lunes con parada 108: se descuenta también del TP de la franja afectada
        #   - Resto: hora completa -> paga 1.00 h
        meal_fixed_sec = 0
        if int(hh) == 7:
            meal_fixed_sec = 15 * 60
        elif int(hh) == 8:
            meal_fixed_sec = 15 * 60
        elif int(hh) == 13:
            meal_fixed_sec = 30 * 60

        # Descuento exacto: fijo + mantenimiento 108 de lunes.
        meal_applied_sec = max(0, min(3600, meal_fixed_sec + max(0, maint108_sec)))
        tp_h_excel = round((3600 - meal_applied_sec) / 3600.0, 3)

        # ✅ Si el Tiempo Trabajado (TW = TP - TX) sale negativo, se corrige el TX
        # restándole el DOBLE del TW negativo. Así TW pasa a ser positivo.
        #   Ej: TP=0.75, TX=0.90 -> TW=-0.15
        #       TX' = 0.90 - 2*0.15 = 0.60 -> TW' = 0.75 - 0.60 = +0.15
        paid_sec = int(round(tp_h_excel * 3600))
        tx_sec = max(0, int(tx_sec))
        # Si el bucket viejo todavía traía 108 dentro de TX, retirarlo para no duplicar
        # el descuento, porque 108 ya fue descontado del TP.
        if maint108_sec > 0 and int(b.get("maint108Sec", 0) or 0) <= 0:
            tx_sec = max(0, tx_sec - maint108_sec)
        tw_sec = paid_sec - tx_sec
        if tw_sec < 0:
            tx_sec = max(0, tx_sec - (2 * abs(tw_sec)))

        tcnp_sec = float(b.get("tcnpSec", 0.0) or 0.0)

        rows.append({
            "fecha": day0.to_pydatetime(),
            "referencia": ", ".join(referencias),
            "equipo": "HP",
            "operario": op_export,
            "inicio": h0.to_pydatetime(),
            "fin": h1.to_pydatetime(),
            "pp": None,
            "pn": int(cut),
            "rx": int(max(0, reject)),
            "tp_h": float(tp_h_excel),
            "tx_h": round(max(0, tx_sec) / 3600.0, 3),
            "tc_h_uni": (tcnp_sec / 3600.0) if tcnp_sec > 0 else None,
        })

    return rows


# =========================================================
# HP: métricas EXACTAS tipo Excel exportado
# =========================================================
def _hp_float(x: Any, default: float = 0.0) -> float:
    try:
        v = float(x if x is not None else default)
        if np.isfinite(v):
            return v
    except Exception:
        pass
    return float(default)


def _hp_excel_ratio(num: float, den: float) -> float:
    den = _hp_float(den)
    if abs(den) <= 1e-12:
        return 0.0
    return _hp_float(num) / den


def _hp_h_to_hhmm(h: float) -> str:
    sec = int(round(max(0.0, _hp_float(h)) * 3600.0))
    return _seconds_to_hhmm(float(sec))


def _hp_pct_text(ratio: float) -> str:
    return f"{(_hp_float(ratio) * 100.0):.1f}%"


def _hp_export_row_metrics(row: dict, *, oee_expected: float = 0.70) -> dict:
    """
    Replica las fórmulas de la plantilla Excel OEE para una fila HP:
      G PP  = OEE esperado * VN * TP
      J PB  = PN - RX
      M TW  = TP - TX
      O VN  = 1 / TC
      P TCN = PN * TC
      Q TCE = PB * TC
      R ID  = TW / TP
      S IEO = TCN / TW
      T IC  = TCE / TCN
      U OEE = TCE / TP
      V PA  = PN / PP
    La plantilla usa horas en K/L/M/N/P/Q; por eso aquí todo el tiempo está en horas.
    """
    pn = _hp_float(row.get("pn"), 0.0)
    rx = _hp_float(row.get("rx"), 0.0)
    pb = max(0.0, pn - rx)

    tp_h = max(0.0, _hp_float(row.get("tp_h"), 0.0))
    tx_h = max(0.0, _hp_float(row.get("tx_h"), 0.0))
    tw_h = tp_h - tx_h

    # En el Excel TW puede quedar negativo si TX supera TP; el generador HP ya corrige TX,
    # pero protegemos el dashboard ante datos viejos o atípicos.
    if tw_h < 0:
        tx_h = max(0.0, tx_h - (2.0 * abs(tw_h)))
        tw_h = tp_h - tx_h
    tw_h = max(0.0, tw_h)

    tc_h = max(0.0, _hp_float(row.get("tc_h_uni"), 0.0))
    vn = (1.0 / tc_h) if tc_h > 0 else 0.0

    try:
        meta_factor = float(row.get("meta_factor", 1.0) or 1.0)
    except Exception:
        meta_factor = 1.0
    if not np.isfinite(meta_factor) or meta_factor <= 0:
        meta_factor = 1.0

    pp = max(0.0, _hp_float(oee_expected, 0.70) * vn * tp_h * meta_factor)
    tcn_h = pn * tc_h
    tce_h = pb * tc_h

    return {
        "pp": float(pp),
        "pn": float(pn),
        "rx": float(rx),
        "pb": float(pb),
        "tp_h": float(tp_h),
        "tx_h": float(tx_h),
        "tw_h": float(tw_h),
        "tc_h": float(tc_h),
        "vn": float(vn),
        "tcn_h": float(tcn_h),
        "tce_h": float(tce_h),
        "id": _hp_excel_ratio(tw_h, tp_h),
        "ieo": _hp_excel_ratio(tcn_h, tw_h),
        "ic": _hp_excel_ratio(tce_h, tcn_h),
        "oee": _hp_excel_ratio(tce_h, tp_h),
        "pa": _hp_excel_ratio(pn, pp),
    }


def _hp_export_summary_metrics(rows: list[dict], *, oee_expected: float = 0.70) -> dict:
    mets = [_hp_export_row_metrics(r, oee_expected=oee_expected) for r in (rows or [])]

    def total(k: str) -> float:
        return float(sum(_hp_float(m.get(k), 0.0) for m in mets))

    out = {
        "pp": total("pp"),
        "pn": total("pn"),
        "rx": total("rx"),
        "pb": total("pb"),
        "tp_h": total("tp_h"),
        "tx_h": total("tx_h"),
        "tw_h": total("tw_h"),
        "tc_h_sum": total("tc_h"),
        "tcn_h": total("tcn_h"),
        "tce_h": total("tce_h"),
    }

    out["id"] = _hp_excel_ratio(out["tw_h"], out["tp_h"])
    out["ieo"] = _hp_excel_ratio(out["tcn_h"], out["tw_h"])
    out["ic"] = _hp_excel_ratio(out["tce_h"], out["tcn_h"])
    out["oee"] = _hp_excel_ratio(out["tce_h"], out["tp_h"])
    out["pa"] = _hp_excel_ratio(out["pn"], out["pp"])
    return out


def _hp_oee_chart_from_export_rows(rows: list[dict], *, oee_expected: float = 0.70) -> dict:
    labels, oee, operacional, disponibilidad, calidad = [], [], [], [], []
    for row in rows or []:
        try:
            ts = pd.Timestamp(row.get("inicio"))
            label = f"{int(ts.hour):02d}:00"
        except Exception:
            label = str(row.get("inicio") or "").strip()
        if not label:
            continue

        m = _hp_export_row_metrics(row, oee_expected=oee_expected)
        labels.append(label)
        oee.append(round(m["oee"] * 100.0, 1))
        operacional.append(round(m["ieo"] * 100.0, 1))
        disponibilidad.append(round(m["id"] * 100.0, 1))
        calidad.append(round(m["ic"] * 100.0, 1))

    return {
        "labels": labels,
        "oee": oee,
        "operacional": operacional,
        "disponibilidad": disponibilidad,
        "calidad": calidad,
    }


def _seconds_to_decimal_hours(sec: float) -> str:
    """
    Convierte segundos a horas decimales con 2 decimales.
    Ejemplo: 1800 segundos -> "0.50" (30 minutos)
    """
    try:
        h = float(sec or 0.0) / 3600.0
        return f"{h:.2f}"
    except Exception:
        return "0.00"


def _hp_apply_excel_metrics_to_result(result: dict, day_exports: list[dict], *, oee_expected: float = 0.70) -> dict:
    """Hace que los KPIs del dashboard HP sean los mismos totales del Excel descargable."""
    if not isinstance(result, dict):
        return result

    all_rows: list[dict] = []
    for item in day_exports or []:
        all_rows.extend(item.get("rows", []) or [])
    if not all_rows:
        return result

    summary = _hp_export_summary_metrics(all_rows, oee_expected=oee_expected)

    kpis = result.setdefault("kpis", {})
    ui = result.setdefault("kpis_ui", {})

    # Guardar trazabilidad completa del Excel para depuración / frontend.
    kpis["hp_excel_summary"] = summary
    kpis["oee_esperado"] = float(oee_expected)

    # Totales equivalentes a la fila TOTAL del Excel.
    kpis["oee"] = float(summary["oee"])
    kpis["oee_rendimiento"] = float(summary["ieo"])
    kpis["oee_disponibilidad"] = float(summary["id"])
    kpis["oee_calidad"] = float(summary["ic"])
    kpis["cumplimiento_plan"] = float(summary["pa"])
    kpis["produccion_planeada_calc"] = float(summary["pp"])

    kpis["tiempo_pagado_h"] = float(summary["tp_h"])
    kpis["tx_h"] = float(summary["tx_h"])
    kpis["tw_h"] = float(summary["tw_h"])
    kpis["tiempo_trabajado_h"] = float(summary["tw_h"])

    kpis["horas_disponibles_sec"] = int(round(summary["tp_h"] * 3600.0))
    kpis["otro_tiempo_sec"] = int(round(summary["tx_h"] * 3600.0))
    kpis["tiempo_efectivo_sec"] = int(round(summary["tw_h"] * 3600.0))

    kpis["horas_disponibles_hhmm"] = _hp_h_to_hhmm(summary["tp_h"])
    kpis["otro_tiempo_hhmm"] = _hp_h_to_hhmm(summary["tx_h"])
    kpis["tiempo_efectivo_hhmm"] = _hp_h_to_hhmm(summary["tw_h"])

    ui["OEE"] = (
        f"{_hp_pct_text(summary['oee'])}||"
        f"Indice de Eficiencia Operacional: {_hp_pct_text(summary['ieo'])} | "
        f"Indice de Disponibilidad: {_hp_pct_text(summary['id'])} | "
        f"Indice de Calidad: {_hp_pct_text(summary['ic'])}"
    )
    ui["Cumplimiento del plan"] = _hp_pct_text(summary["pa"])
    ui["Tiempo Pagado"] = _hp_h_to_hhmm(summary["tp_h"])
    ui["Tiempo Trabajado"] = f"{_hp_h_to_hhmm(summary['tw_h'])} ({_hp_pct_text(summary['id'])})"

    # ✅ Calcular totales de paradas desde buckets/day_exports
    comida_sec = 0
    pausa_sec = 0
    no_reg_sec = 0
    for item in day_exports or []:
        for bucket in item.get("buckets", []) or []:
            comida_sec += int(bucket.get("mealSec", 0) or 0)
            pausa_sec += int(bucket.get("pauseSec", 0) or 0)
            no_reg_sec += int(bucket.get("unregSec", 0) or 0)  # ✅ CORRECTO: "unregSec" no "noRegistradoSec"

    tx_pct = _hp_excel_ratio(summary["tx_h"], summary["tp_h"])

    # ✅ Generar KPI con desglose (formato THB) - IGUAL ORDEN QUE THB
    ui["Tiempos Perdidos"] = (
        f"{_seconds_to_decimal_hours(kpis['otro_tiempo_sec'])}"
        f"||105: {_seconds_to_decimal_hours(comida_sec)}"
        f" | No registrado: {_seconds_to_decimal_hours(no_reg_sec)}"
        f" | 104: {_seconds_to_decimal_hours(pausa_sec)}"
        f" | {_hp_pct_text(tx_pct)}"
    )

    # Enriquecer las tarjetas por hora DAY con los mismos datos de las filas del Excel.
    # El frontend ya prioriza tpH/twH/txH/oeePct/paPct cuando existen.
    row_by_hour: dict[str, dict] = {}
    for row in all_rows:
        try:
            hkey = f"{int(pd.Timestamp(row.get('inicio')).hour):02d}"
        except Exception:
            continue
        row_by_hour[hkey] = row

    buckets = kpis.get("hourly_buckets")
    if isinstance(buckets, list) and buckets:
        for b in buckets:
            raw_h = str(b.get("hour") or b.get("hourLabel") or "").strip()
            m_h = re.search(r"\d{1,2}", raw_h)
            hkey = f"{int(m_h.group(0)):02d}" if m_h else ""
            row = row_by_hour.get(hkey)
            if not row:
                continue
            met = _hp_export_row_metrics(row, oee_expected=oee_expected)
            b["tpH"] = round(met["tp_h"], 6)
            b["txH"] = round(met["tx_h"], 6)
            b["twH"] = round(met["tw_h"], 6)
            b["idPct"] = round(met["id"] * 100.0, 4)
            b["oeePct"] = round(met["oee"] * 100.0, 4)
            b["paPct"] = round(met["pa"] * 100.0, 4)
            b["plan"] = round(met["pp"], 6)
            b["planned"] = round(met["pp"], 6)
            b["excel_metrics"] = met

    if str(result.get("period") or "").strip().lower() == "day":
        result["oee_chart"] = _hp_oee_chart_from_export_rows(all_rows, oee_expected=oee_expected)

    return result


def _build_hp_export_day_exports(
        *,
        df_corte: pd.DataFrame,
        df_verif: pd.DataFrame,
        df_paradas: pd.DataFrame,
        df_duracion: pd.DataFrame,
        period: str,
        period_value: str,
        operator: str = "",
        time_start: str = "",
        time_end: str = "",
        time_apply: Any = False,
        base_result: Optional[dict] = None,
) -> list[dict]:
    """
    Construye las mismas filas que se envían al Excel HP.
    Se usa tanto para descargar el Excel como para alimentar el dashboard.
    """
    days = _hp_days_for_period_export(df_corte, period, period_value)
    if not days:
        return []

    day_exports: list[dict] = []
    p = str(period or "day").strip().lower()
    base_day = None
    try:
        if p == "day" and base_result and base_result.get("status") == "ok":
            base_day = pd.Timestamp(period_value).normalize()
    except Exception:
        base_day = None

    for day in days:
        day = pd.Timestamp(day).normalize()

        if base_day is not None and day == base_day:
            result_day = base_result
        else:
            result_day = analyze_hp(
                df_corte,
                df_verif,
                period="day",
                period_value=day.strftime("%Y-%m-%d"),
                machine="HP",
                operator=operator,
                time_start=(time_start if time_apply else ""),
                time_end=(time_end if time_apply else ""),
                time_apply=time_apply,
                df_paradas=df_paradas,
                df_duracion=df_duracion,
            )

        if not isinstance(result_day, dict) or result_day.get("status") != "ok":
            continue

        rows = _hp_export_rows_for_day(
            df_corte=df_corte,
            day=day,
            buckets=result_day.get("kpis", {}).get("hourly_buckets", []),
            operator=operator,
        )
        if not rows:
            continue

        day_exports.append({
            "sheet_title": day.strftime("%Y-%m-%d"),
            "rows": rows,
            "buckets": result_day.get("kpis", {}).get("hourly_buckets", []),  # ✅ AGREGAR BUCKETS
            "oee_expected": float(0.70),
        })

    return day_exports


def _safe_xlsx_sheet_title(title: str) -> str:
    s = str(title or "").strip()
    s = re.sub(r'[\\/*?:\[\]]+', "-", s)
    s = s[:31].strip()
    return s or "Hoja"


def _get_thb_export_template_path() -> str:
    """
    Busca el modelo OEE para la exportación THB.
    Prioridad:
      1) misma carpeta de app.py
      2) subcarpeta templates_excel/
      3) uploads/
      4) rutas de trabajo locales (útil en desarrollo)
    """
    base_dir = Path(__file__).resolve().parent
    candidates = [
        base_dir / "Modelo OEE.xlsx",
        base_dir / "Modelo OEE - Copia.xlsx",
        base_dir / "templates_excel" / "Modelo OEE.xlsx",
        base_dir / "templates_excel" / "Modelo OEE - Copia.xlsx",
        base_dir / app.config.get("UPLOAD_FOLDER", "uploads") / "Modelo OEE.xlsx",
        base_dir / app.config.get("UPLOAD_FOLDER", "uploads") / "Modelo OEE - Copia.xlsx",
        Path("/mnt/data/Modelo OEE.xlsx"),
        Path("/mnt/data/Modelo OEE - Copia.xlsx"),
    ]

    for p in candidates:
        try:
            if p.exists():
                return str(p)
        except Exception:
            continue

    raise FileNotFoundError(
        "No encontré el archivo 'Modelo OEE.xlsx'. Déjalo en la misma carpeta de app.py o en templates_excel/."
    )


def _thb_days_for_period_export(df_verif: pd.DataFrame, period: str, period_value: str, operator: str = "") -> list[
    pd.Timestamp]:
    cols = _resolve_thb_cols(df_verif)
    if not cols.col_fecha:
        return []

    dt = _thb_build_dt(df_verif, cols.col_fecha, cols.col_hora)
    pv = _thb_normalize_period_value(period, period_value)
    r0, r1 = _thb_period_bounds(period, pv)
    if r0 is None or r1 is None:
        return []

    m = dt.notna() & (dt >= r0) & (dt < r1)

    op = (operator or "").strip()
    if op and op.lower() != "general" and cols.col_nombre and cols.col_nombre in df_verif.columns:
        m = m & df_verif[cols.col_nombre].astype(str).str.strip().eq(op)

    days = (
        pd.to_datetime(dt.loc[m], errors="coerce")
        .dropna()
        .dt.normalize()
        .drop_duplicates()
        .sort_values()
        .tolist()
    )
    return [pd.Timestamp(d) for d in days]


def _thb_paid_hours_for_export_row(hour_int: int) -> float:
    """
    Tiempo pagado por hora según la regla solicitada:
      - 07:00 a 08:00 -> 0.75 h (entrada 07:15)
      - 08:00 a 09:00 -> 0.75 h
      - 13:00 a 14:00 -> 0.50 h
      - resto         -> 1.00 h
    """
    try:
        hh = int(hour_int)
    except Exception:
        hh = -1

    if hh == 7:
        return 0.75
    if hh == 8:
        return 0.75
    if hh == 13:
        return 0.50
    return 1.00


ARNES_TO_REFERENCIA = {
    "50000120": "RIDUCO",
    "50010521": "SOFTECON",
    "50020120": "RIDUCO",
    "50020425": "CHR",
    "50020521": "SOFTECON",
    "50030521": "SOFTECON",
    "50040521": "SOFTECON",
    "50050521": "SOFTECON",
    "50060820": "CR4 - 160",
    "50070521": "SOFTECON",
    "50080521": "SOFTECON",
    "50090521": "SOFTECON",
    "50130625": "CR4",
    "50211025": "NKD",
    "50221125": "CR4 - NUEVA",
    "50231025": "CHR",
    "50301021": "SPECIAL",
    "50001221": "HERO",
}


def _clean_arnes_code(x: object) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    if not s or s.lower() in ("nan", "none"):
        return ""
    s = re.sub(r"\.0$", "", s)
    s = re.sub(r"\s+", "", s)
    return s


def _map_referencia_from_arnes(codigo_arnes: object) -> str:
    code = _clean_arnes_code(codigo_arnes)
    return ARNES_TO_REFERENCIA.get(code, "")


def _thb_export_rows_for_day(
        df_verif: pd.DataFrame,
        day: pd.Timestamp,
        buckets: list[dict],
        operator: str = "",
        df_paradas: pd.DataFrame | None = None,
) -> list[dict]:
    cols = _resolve_thb_cols(df_verif)
    if not cols.col_fecha:
        return []

    dt = _thb_build_dt(df_verif, cols.col_fecha, cols.col_hora)
    day0 = pd.Timestamp(day).normalize()
    day1 = day0 + pd.Timedelta(days=1)

    m = dt.notna() & (dt >= day0) & (dt < day1)

    op_fixed = (operator or "").strip()
    if op_fixed and op_fixed.lower() != "general" and cols.col_nombre and cols.col_nombre in df_verif.columns:
        m = m & df_verif[cols.col_nombre].astype(str).str.strip().eq(op_fixed)

    sub = df_verif.loc[m].copy()
    dt_sub = dt.loc[m].copy()

    c_total = cols.col_total_piezas if cols.col_total_piezas in sub.columns else None
    c_buenas = cols.col_piezas_buenas if cols.col_piezas_buenas in sub.columns else None
    c_nombre = cols.col_nombre if cols.col_nombre in sub.columns else None
    c_arnes = cols.col_arnes if cols.col_arnes in sub.columns else None

    bucket_by_hour = {}
    for b in (buckets or []):
        try:
            hh = int(str(b.get("hour") or "0"))
        except Exception:
            continue
        bucket_by_hour[hh] = b

    def _thb_clean_stop_code(x: object) -> str:
        s = str(x or "").strip()
        if not s:
            return ""
        m_code = re.search(r"\d+", s)
        if not m_code:
            return ""
        try:
            return str(int(m_code.group(0)))
        except Exception:
            return m_code.group(0)

    def _thb_bucket_has_stop_108(bucket: dict) -> bool:
        if not isinstance(bucket, dict):
            return False

        # Soporta diferentes nombres si el bucket ya trae el 108 separado.
        for k in ("maint108Sec", "tpDiscount108Sec", "parada108Sec", "stop108Sec", "p108Sec", "108Sec"):
            try:
                if float(bucket.get(k, 0) or 0) > 0:
                    return True
            except Exception:
                pass

        # Soporta detalle visual de paradas: other_causes / causes / paradas / stops.
        for list_key in ("other_causes", "causes", "paradas", "stops", "items"):
            items = bucket.get(list_key, [])
            if not isinstance(items, list):
                continue

            for item in items:
                if isinstance(item, dict):
                    code = (
                        item.get("code") or item.get("codigo") or item.get("código") or
                        item.get("causa") or item.get("label") or item.get("id")
                    )
                else:
                    code = item

                if _thb_clean_stop_code(code) == "108":
                    return True

        return False

    def _thb_duration_to_seconds_local(values: pd.Series) -> pd.Series:
        if values is None:
            return pd.Series([], dtype="float64")

        if pd.api.types.is_numeric_dtype(values):
            x = pd.to_numeric(values, errors="coerce").fillna(0.0).astype(float)
            nz = x[x > 0].dropna()
            if nz.empty:
                return x

            mx = float(nz.max())
            med = float(nz.median())

            # Fracción de día Excel.
            if mx <= 2.0:
                return x * 86400.0

            # En Sheets suele venir en minutos cuando todos los valores son pequeños.
            if mx <= 240.0 and med <= 60.0:
                return x * 60.0

            return x

        ss = values.astype(str).str.replace("\u00a0", " ", regex=False).str.strip()
        td = pd.to_timedelta(ss, errors="coerce")
        sec = td.dt.total_seconds()

        m = sec.isna()
        if m.any():
            nums = pd.to_numeric(ss.where(m).str.replace(",", ".", regex=False), errors="coerce")
            if nums.notna().any():
                sec = sec.fillna(_thb_duration_to_seconds_local(nums))

        return pd.to_numeric(sec, errors="coerce").fillna(0.0).clip(lower=0.0).astype(float)

    def _thb_registered_108_hours_for_day(df_paradas_src: pd.DataFrame | None) -> set[int]:
        hours_108: set[int] = set()
        if df_paradas_src is None or not isinstance(df_paradas_src, pd.DataFrame) or df_paradas_src.empty:
            return hours_108

        dfp = df_paradas_src.copy()
        c_fecha_p = _find_col(dfp, "fecha")
        c_hora_p = _find_col(dfp, "hora")
        c_codigo_p = (
            _find_col(dfp, "código") or _find_col(dfp, "codigo") or
            _find_col(dfp, "cod") or _find_col(dfp, "causa") or
            _find_col(dfp, "parada") or _find_col(dfp, "motivo")
        )

        if not c_fecha_p or not c_hora_p or not c_codigo_p or c_codigo_p not in dfp.columns:
            return hours_108

        dt_end_p = pd.to_datetime(_build_dt_from_fecha_hora(dfp, c_fecha_p, c_hora_p), errors="coerce")
        codes_p = dfp[c_codigo_p].apply(_thb_clean_stop_code)
        m108 = codes_p.eq("108") & dt_end_p.notna()
        if not m108.any():
            return hours_108

        dt_108 = dt_end_p.loc[m108]
        m_day_108 = dt_108.dt.normalize().eq(day0)
        dt_108 = dt_108.loc[m_day_108]
        if dt_108.empty:
            return hours_108

        # Fallback: si la parada 108 quedó registrada con Hora dentro de la franja.
        for h in dt_108.dt.hour.dropna().astype(int).tolist():
            if 0 <= int(h) <= 23:
                hours_108.add(int(h))

        # Si hay Duración, también se marca cualquier hora con la que la parada se solape.
        c_dur_p = None
        for c in dfp.columns:
            k = str(c or "").strip().lower()
            if any(tok in k for tok in ("dur", "tiempo", "min", "seg", "sec")):
                if c not in (c_fecha_p, c_hora_p, c_codigo_p):
                    c_dur_p = c
                    break

        if c_dur_p and c_dur_p in dfp.columns:
            dur_sec = _thb_duration_to_seconds_local(dfp.loc[m108, c_dur_p])
            dur_sec = dur_sec.loc[dt_108.index].fillna(0.0).clip(lower=0.0)
            dt_start_p = dt_108 - pd.to_timedelta(dur_sec, unit="s")

            for start_p, end_p in zip(dt_start_p.tolist(), dt_108.tolist()):
                try:
                    if pd.isna(start_p) or pd.isna(end_p):
                        continue
                    start_p = pd.Timestamp(start_p)
                    end_p = pd.Timestamp(end_p)
                    if end_p <= start_p:
                        continue

                    for hh2 in range(24):
                        h0_2 = day0 + pd.Timedelta(hours=hh2)
                        h1_2 = h0_2 + pd.Timedelta(hours=1)
                        if min(end_p, h1_2) > max(start_p, h0_2):
                            hours_108.add(hh2)
                except Exception:
                    continue

        return hours_108

    thb_108_hours = _thb_registered_108_hours_for_day(df_paradas)

    hours = sorted(bucket_by_hour.keys())
    if not hours and not sub.empty:
        hours = sorted(pd.to_datetime(dt_sub, errors="coerce").dropna().dt.hour.unique().tolist())

    rows = []
    last_operator = op_fixed if (op_fixed and op_fixed.lower() != "general") else "General"

    for hh in hours:
        h0 = day0 + pd.Timedelta(hours=int(hh))
        h1 = h0 + pd.Timedelta(hours=1)

        dt_hour = pd.to_datetime(dt_sub, errors="coerce")
        in_hour = dt_hour.notna() & (dt_hour >= h0) & (dt_hour < h1)
        hour_df = sub.loc[in_hour].copy()

        planned_calc = 0
        cut = 0

        if c_total:
            tot = pd.to_numeric(hour_df[c_total], errors="coerce").fillna(0).astype(float)
        else:
            tot = pd.Series([], dtype="float64")

        if c_buenas:
            buenas = pd.to_numeric(hour_df[c_buenas], errors="coerce").fillna(0).astype(float)
        else:
            buenas = pd.Series([], dtype="float64")

        if not tot.empty:
            if not buenas.empty:
                planned_calc = int(tot.where(tot > 0, buenas).sum())
            else:
                planned_calc = int(tot.sum())

        if not buenas.empty:
            cut = int(buenas.sum())
        elif not tot.empty:
            cut = int(tot.sum())

        # No conformes = cortados - planeados
        reject = max(0, int(cut) - int(planned_calc))

        if c_nombre and not hour_df.empty and not op_fixed:
            ops = hour_df[c_nombre].dropna().astype(str).str.strip()
            ops = ops[ops.ne("")]
            if not ops.empty:
                try:
                    last_operator = ops.mode().iloc[0]
                except Exception:
                    last_operator = ops.iloc[0]

        b = bucket_by_hour.get(int(hh), {})

        # =========================
        # EXCEL THB
        # =========================
        # TX real del bucket (SIN comida)
        tx_sec = int(b.get("otherDeadSec", 0) or 0)

        # comida no paga del bucket
        meal_sec = int(b.get("mealSec", 0) or 0)

        # comida fija por hora
        # Entrada del turno 07:15 -> la hora 7 tiene 15 min no trabajados (7:00-7:15).
        meal_fixed_sec = 0
        if int(hh) == 7:
            meal_fixed_sec = 15 * 60
        elif int(hh) == 8:
            meal_fixed_sec = 15 * 60
        elif int(hh) == 13:
            meal_fixed_sec = 30 * 60

        # si hay 105 mayor al fijo, se descuenta la real;
        # si es menor, se conserva el fijo
        meal_applied_sec = max(meal_fixed_sec, meal_sec) if meal_fixed_sec > 0 else 0

        # THB - Descargar Excel:
        # Los lunes, SOLO en 07:00-08:00, si la parada 108 existe registrada,
        # se descuentan 30 minutos adicionales del Tiempo Pagado.
        # El TX permanece igual: conserva las paradas registradas normales.
        has_thb_108_lunes_7 = (
            int(hh) == 7
            and int(day0.weekday()) == 0
            and ((7 in thb_108_hours) or _thb_bucket_has_stop_108(b))
        )
        if has_thb_108_lunes_7:
            meal_applied_sec += 30 * 60

        meal_applied_sec = max(0, min(3600, meal_applied_sec))

        # TP del Excel:
        # - horas normales: 1.00
        # - desayuno/almuerzo: 1h menos comida aplicada
        tp_h_excel = round((3600 - meal_applied_sec) / 3600.0, 3)

        # si no hay producción buena pero sí hay tiempo perdido,
        # TP sigue siendo la hora pagada de esa franja
        # (1.00 normal, 0.75 desayuno, 0.50 almuerzo, o menos si 105 > fijo)
        if int(cut) <= 0 and int(tx_sec) > 0:
            tp_h_excel = round((3600 - meal_applied_sec) / 3600.0, 3)

        # TX del Excel NO suma comida
        lost_sec = int(tx_sec)

        # Tiempo de ciclo nominal ponderado (h/uni)
        tcnp_sec = float(b.get("tcnpSec", 0.0) or 0.0)

        referencias = []

        if c_arnes and c_arnes in hour_df.columns and not hour_df.empty:
            vals_arnes = hour_df[c_arnes].dropna().astype(str).str.strip()
            vals_arnes = vals_arnes[vals_arnes.ne("")]

            vistos = set()
            for arnes_raw in vals_arnes.tolist():
                codigo_arnes = _clean_arnes_code(arnes_raw)
                if not codigo_arnes:
                    continue

                ref = _map_referencia_from_arnes(codigo_arnes)
                if not ref:
                    continue

                ref_key = ref.strip().upper()
                if ref_key in vistos:
                    continue

                vistos.add(ref_key)
                referencias.append(ref.strip())

        referencia = ", ".join(referencias)

        rows.append({
            "fecha": day0.to_pydatetime(),
            "referencia": referencia,
            "equipo": "THB",
            "operario": (op_fixed if (op_fixed and op_fixed.lower() != "general") else last_operator) or "General",
            "inicio": h0.to_pydatetime(),
            "fin": h1.to_pydatetime(),
            "pp": None,  # Producción planeada vacía
            "pn": int(cut),  # si no hubo producción buena, queda 0
            "rx": int(reject),
            "tp_h": float(tp_h_excel),
            "tx_h": round(lost_sec / 3600.0, 3),
            "tc_h_uni": (tcnp_sec / 3600.0) if tcnp_sec > 0 else None,
        })

    return rows



# =========================================================
# THB: métricas EXACTAS tipo Excel exportado
# =========================================================
def _thb_export_float(x: Any, default: float = 0.0) -> float:
    try:
        v = float(x if x is not None else default)
        if np.isfinite(v):
            return v
    except Exception:
        pass
    return float(default)


def _thb_export_ratio(num: float, den: float) -> float:
    den = _thb_export_float(den)
    if abs(den) <= 1e-12:
        return 0.0
    return _thb_export_float(num) / den


def _thb_export_h_to_hhmm(h: float) -> str:
    sign = "-" if _thb_export_float(h) < 0 else ""
    sec = int(round(abs(_thb_export_float(h)) * 3600.0))
    return sign + _seconds_to_hhmm(float(sec))


def _thb_export_pct_text(ratio: float) -> str:
    return f"{(_thb_export_float(ratio) * 100.0):.1f}%"


def _thb_export_oee_expected(default: float = 0.70) -> float:
    """
    Lee el OEE esperado que usa el modelo Excel THB (celda G2).
    Si no puede leerlo, conserva 70%, que es el valor histórico usado por THB.
    """
    try:
        from openpyxl import load_workbook

        template_path = _get_thb_export_template_path()
        wb = load_workbook(template_path, read_only=True, data_only=True)
        ws = wb["Por hora"] if "Por hora" in wb.sheetnames else wb.active
        raw = ws["G2"].value
        wb.close()

        val = _thb_export_float(raw, default)
        if val > 1.5 and val <= 100.0:
            val = val / 100.0
        if val > 0:
            return float(val)
    except Exception:
        pass
    return float(default)


def _thb_export_row_metrics(row: dict, *, oee_expected: float = 0.70) -> dict:
    """
    Replica las fórmulas de la plantilla Excel OEE para una fila THB:
      G PP  = OEE esperado * VN * TP
      J PB  = PN - RX
      M TW  = TP - TX
      O VN  = 1 / TC
      P TCN = PN * TC
      Q TCE = PB * TC
      R ID  = TW / TP
      S IEO = TCN / TW
      T IC  = TCE / TCN
      U OEE = TCE / TP
      V PA  = PN / PP
    """
    pn = _thb_export_float(row.get("pn"), 0.0)
    rx = _thb_export_float(row.get("rx"), 0.0)
    pb = pn - rx

    tp_h = _thb_export_float(row.get("tp_h"), 0.0)
    tx_h = _thb_export_float(row.get("tx_h"), 0.0)
    tw_h = tp_h - tx_h

    tc_h = max(0.0, _thb_export_float(row.get("tc_h_uni"), 0.0))
    vn = (1.0 / tc_h) if tc_h > 0 else 0.0

    try:
        meta_factor = float(row.get("meta_factor", 1.0) or 1.0)
    except Exception:
        meta_factor = 1.0
    if not np.isfinite(meta_factor) or meta_factor <= 0:
        meta_factor = 1.0

    pp = max(0.0, _thb_export_float(oee_expected, 0.70) * vn * tp_h * meta_factor)
    tcn_h = pn * tc_h
    tce_h = pb * tc_h

    return {
        "pp": float(pp),
        "pn": float(pn),
        "rx": float(rx),
        "pb": float(pb),
        "tp_h": float(tp_h),
        "tx_h": float(tx_h),
        "tw_h": float(tw_h),
        "tc_h": float(tc_h),
        "vn": float(vn),
        "tcn_h": float(tcn_h),
        "tce_h": float(tce_h),
        "id": _thb_export_ratio(tw_h, tp_h),
        "ieo": _thb_export_ratio(tcn_h, tw_h),
        "ic": _thb_export_ratio(tce_h, tcn_h),
        "oee": _thb_export_ratio(tce_h, tp_h),
        "pa": _thb_export_ratio(pn, pp),
    }


def _thb_export_summary_metrics(rows: list[dict], *, oee_expected: float = 0.70) -> dict:
    mets = [_thb_export_row_metrics(r, oee_expected=oee_expected) for r in (rows or [])]

    def total(k: str) -> float:
        return float(sum(_thb_export_float(m.get(k), 0.0) for m in mets))

    out = {
        "pp": total("pp"),
        "pn": total("pn"),
        "rx": total("rx"),
        "pb": total("pb"),
        "tp_h": total("tp_h"),
        "tx_h": total("tx_h"),
        "tw_h": total("tw_h"),
        "tc_h_sum": total("tc_h"),
        "tcn_h": total("tcn_h"),
        "tce_h": total("tce_h"),
    }

    out["id"] = _thb_export_ratio(out["tw_h"], out["tp_h"])
    out["ieo"] = _thb_export_ratio(out["tcn_h"], out["tw_h"])
    out["ic"] = _thb_export_ratio(out["tce_h"], out["tcn_h"])
    out["oee"] = _thb_export_ratio(out["tce_h"], out["tp_h"])
    out["pa"] = _thb_export_ratio(out["pn"], out["pp"])
    return out


def _thb_oee_chart_from_export_rows(rows: list[dict], *, oee_expected: float = 0.70) -> dict:
    labels, oee, operacional, disponibilidad, calidad = [], [], [], [], []
    for row in rows or []:
        try:
            ts = pd.Timestamp(row.get("inicio"))
            label = f"{int(ts.hour):02d}:00"
        except Exception:
            label = str(row.get("inicio") or "").strip()
        if not label:
            continue

        m = _thb_export_row_metrics(row, oee_expected=oee_expected)
        labels.append(label)
        oee.append(round(m["oee"] * 100.0, 1))
        operacional.append(round(m["ieo"] * 100.0, 1))
        disponibilidad.append(round(m["id"] * 100.0, 1))
        calidad.append(round(m["ic"] * 100.0, 1))

    return {
        "labels": labels,
        "oee": oee,
        "operacional": operacional,
        "disponibilidad": disponibilidad,
        "calidad": calidad,
    }


def _build_thb_export_day_exports(
        *,
        df_verif: pd.DataFrame,
        df_paradas: pd.DataFrame,
        period: str,
        period_value: str,
        operator: str = "",
        base_result: Optional[dict] = None,
) -> list[dict]:
    """
    Construye las mismas filas que se envían al Excel THB.
    Se usa para que el dashboard THB pueda reflejar exactamente el Excel descargable.
    """
    days = _thb_days_for_period_export(df_verif, period, period_value, operator=operator)
    if not days:
        return []

    day_exports: list[dict] = []
    p = str(period or "day").strip().lower()
    base_day = None
    try:
        if p == "day" and base_result and base_result.get("status") == "ok":
            base_day = pd.Timestamp(period_value).normalize()
    except Exception:
        base_day = None

    for day in days:
        day = pd.Timestamp(day).normalize()

        if base_day is not None and day == base_day:
            result_day = base_result
        else:
            result_day = analyze_thb(
                df_verif,
                period="day",
                period_value=day.strftime("%Y-%m-%d"),
                machine="THB",
                operator=operator,
                time_start="",
                time_end="",
                df_paradas=df_paradas,
            )

        if not isinstance(result_day, dict) or result_day.get("status") != "ok":
            continue

        buckets = result_day.get("kpis", {}).get("hourly_buckets", [])
        rows = _thb_export_rows_for_day(
            df_verif=df_verif,
            day=day,
            buckets=buckets,
            operator=operator,
        )
        if not rows:
            continue

        day_exports.append({
            "sheet_title": day.strftime("%Y-%m-%d"),
            "rows": rows,
            "buckets": buckets,
        })

    return day_exports


def _thb_apply_excel_metrics_to_result(result: dict, day_exports: list[dict], *, oee_expected: float = 0.70) -> dict:
    """Hace que los KPIs y tarjetas THB sean los mismos totales del Excel descargable."""
    if not isinstance(result, dict):
        return result

    all_rows: list[dict] = []
    for item in day_exports or []:
        all_rows.extend(item.get("rows", []) or [])
    if not all_rows:
        return result

    summary = _thb_export_summary_metrics(all_rows, oee_expected=oee_expected)

    kpis = result.setdefault("kpis", {})
    ui = result.setdefault("kpis_ui", {})

    # Totales equivalentes a la fila TOTAL del Excel.
    kpis["thb_excel_summary"] = summary
    kpis["oee_esperado"] = float(oee_expected)
    kpis["produccion_planeada_calc"] = float(summary["pp"])
    kpis["cumplimiento_plan"] = float(summary["pa"])
    kpis["oee"] = float(summary["oee"])
    kpis["oee_rendimiento"] = float(summary["ieo"])
    kpis["oee_disponibilidad"] = float(summary["id"])
    kpis["oee_calidad"] = float(summary["ic"])

    kpis["circuitos_cortados"] = int(round(summary["pn"]))
    kpis["circuitos_planeados"] = int(round(summary["pb"]))
    kpis["circuitos_no_conformes"] = int(round(summary["rx"]))

    kpis["horas_hombre_sec"] = int(round(summary["tp_h"] * 3600.0))
    kpis["otro_tiempo_sec"] = int(round(summary["tx_h"] * 3600.0))
    kpis["tiempo_efectivo_sec"] = int(round(summary["tw_h"] * 3600.0))

    kpis["horas_hombre_hhmm"] = _thb_export_h_to_hhmm(summary["tp_h"])
    kpis["otro_tiempo_hhmm"] = _thb_export_h_to_hhmm(summary["tx_h"])
    kpis["tiempo_efectivo_hhmm"] = _thb_export_h_to_hhmm(summary["tw_h"])

    ui["Producción Total"] = f"{int(round(summary['pn'])):,}".replace(",", ".")
    ui["Producción Buena"] = f"{int(round(summary['pb'])):,}".replace(",", ".")
    ui["Producción con Defectos"] = f"{int(round(summary['rx'])):,}".replace(",", ".")
    ui["Cumplimiento del plan"] = _thb_export_pct_text(summary["pa"])
    ui["Tiempo Pagado"] = _thb_export_h_to_hhmm(summary["tp_h"])
    ui["Tiempo Trabajado"] = f"{_thb_export_h_to_hhmm(summary['tw_h'])} ({_thb_export_pct_text(summary['id'])})"

    # Mantener el desglose visual existente, pero hacer que el total sea el TX del Excel.
    old_lost = str(ui.get("Tiempo Perdido") or ui.get("Tiempos Perdidos") or "")
    detail = ""
    if "||" in old_lost:
        detail = "||" + old_lost.split("||", 1)[1]
    ui["Tiempo Perdido"] = f"{_thb_export_h_to_hhmm(summary['tx_h'])}{detail}"
    ui["Tiempos Perdidos"] = ui["Tiempo Perdido"]

    ui["OEE"] = (
        f"{_thb_export_pct_text(summary['oee'])}||"
        f"Indice de Disponibilidad: {_thb_export_pct_text(summary['id'])} | "
        f"Indice de Eficiencia Operacional: {_thb_export_pct_text(summary['ieo'])} | "
        f"Indice de Calidad: {_thb_export_pct_text(summary['ic'])}"
    )

    # Enriquecer tarjetas por hora con los mismos datos de cada fila del Excel.
    row_by_hour: dict[str, dict] = {}
    for row in all_rows:
        try:
            hkey = f"{int(pd.Timestamp(row.get('inicio')).hour):02d}"
        except Exception:
            continue
        row_by_hour[hkey] = row

    buckets = kpis.get("hourly_buckets")
    if isinstance(buckets, list) and buckets:
        for b in buckets:
            raw_h = str(b.get("hour") or b.get("hourLabel") or "").strip()
            m_h = re.search(r"\d{1,2}", raw_h)
            hkey = f"{int(m_h.group(0)):02d}" if m_h else ""
            row = row_by_hour.get(hkey)
            if not row:
                continue

            met = _thb_export_row_metrics(row, oee_expected=oee_expected)
            b["cut"] = int(round(met["pn"]))
            b["tpH"] = round(met["tp_h"], 6)
            b["txH"] = round(met["tx_h"], 6)
            b["twH"] = round(met["tw_h"], 6)
            b["idPct"] = round(met["id"] * 100.0, 4)
            b["oeePct"] = round(met["oee"] * 100.0, 4)
            b["paPct"] = round(met["pa"] * 100.0, 4)
            b["plan"] = round(met["pp"], 6)
            b["planned"] = round(met["pp"], 6)
            b["produccionPlaneada"] = round(met["pp"], 6)
            b["excel_metrics"] = met

    if str(result.get("period") or "").strip().lower() == "day":
        result["oee_chart"] = _thb_oee_chart_from_export_rows(all_rows, oee_expected=oee_expected)

    return result

def _ensure_thb_template_capacity(ws, needed_rows: int, body_start: int = 7, body_end: int = 38, total_row: int = 40):
    """
    Amplía el cuerpo si alguna vez se necesitaran más filas que las del modelo.
    """
    from copy import copy

    capacity = int(body_end - body_start + 1)
    if needed_rows <= capacity:
        return body_end, total_row

    extra = int(needed_rows - capacity)
    ws.insert_rows(total_row, amount=extra)

    src_row = body_end
    for i in range(extra):
        trg_row = body_end + 1 + i
        for col in range(1, 23):  # A:V
            src = ws.cell(src_row, col)
            dst = ws.cell(trg_row, col)
            if src.has_style:
                dst._style = copy(src._style)
            if src.number_format:
                dst.number_format = src.number_format
            if src.font:
                dst.font = copy(src.font)
            if src.fill:
                dst.fill = copy(src.fill)
            if src.border:
                dst.border = copy(src.border)
            if src.alignment:
                dst.alignment = copy(src.alignment)
            if src.protection:
                dst.protection = copy(src.protection)

        src_h = ws.row_dimensions[src_row].height
        if src_h is not None:
            ws.row_dimensions[trg_row].height = src_h

    return body_end + extra, total_row + extra


def _clear_thb_template_body(ws, body_start: int, body_end: int):
    for row in range(body_start, body_end + 1):
        for col in range(1, 23):  # A:V
            ws.cell(row, col).value = None


def _write_thb_template_totals(ws, body_start: int, body_end: int, total_row: int):
    ws.cell(total_row, 2).value = "TOTAL"

    ws.cell(total_row, 7).value = f'=SUM(G{body_start}:G{body_end})'

    ws.cell(total_row, 8).value = f'=SUM(H{body_start}:H{body_end})'
    ws.cell(total_row, 9).value = f'=SUM(I{body_start}:I{body_end})'
    ws.cell(total_row, 10).value = f'=SUM(J{body_start}:J{body_end})'
    ws.cell(total_row, 11).value = f'=SUM(K{body_start}:K{body_end})'
    ws.cell(total_row, 12).value = f'=SUM(L{body_start}:L{body_end})'
    ws.cell(total_row, 13).value = f'=SUM(M{body_start}:M{body_end})'
    ws.cell(total_row, 14).value = f'=SUM(N{body_start}:N{body_end})'
    ws.cell(total_row, 15).value = None
    ws.cell(total_row, 16).value = f'=SUM(P{body_start}:P{body_end})'
    ws.cell(total_row, 17).value = f'=SUM(Q{body_start}:Q{body_end})'
    ws.cell(total_row, 18).value = f'=IFERROR(M{total_row}/K{total_row},"")'
    ws.cell(total_row, 19).value = f'=IFERROR(P{total_row}/M{total_row},"")'
    ws.cell(total_row, 20).value = f'=IFERROR(Q{total_row}/P{total_row},"")'
    ws.cell(total_row, 21).value = f'=IFERROR(Q{total_row}/K{total_row},"")'
    ws.cell(total_row, 22).value = f'=IFERROR(H{total_row}/G{total_row},"")'




def _thb_productivity_operator_name_from_data(
        df_verif: pd.DataFrame,
        *,
        period: str,
        period_value: str,
        operator: str = "",
) -> str:
    """
    Nombre que se muestra en el título "Productividad por hora THB".
    - Si el usuario filtró una operaria, se usa esa operaria.
    - Si está en General, se toma automáticamente la operaria más frecuente
      del periodo mostrado en Verificación de Medidas.
    Solo alimenta el título del dashboard; no cambia KPIs ni exportación.
    """
    op = str(operator or "").strip()
    if op and op.lower() != "general":
        return op

    try:
        if df_verif is None or not isinstance(df_verif, pd.DataFrame) or df_verif.empty:
            return ""

        cols = _resolve_thb_cols(df_verif)
        if not cols.col_nombre or cols.col_nombre not in df_verif.columns:
            return ""
        if not cols.col_fecha:
            return ""

        dt = _thb_build_dt(df_verif, cols.col_fecha, cols.col_hora)
        pv = _thb_normalize_period_value(period, period_value)
        r0, r1 = _thb_period_bounds(period, pv)

        if r0 is not None and r1 is not None:
            mask = pd.to_datetime(dt, errors="coerce").notna() & (dt >= r0) & (dt < r1)
            sub = df_verif.loc[mask].copy()
        else:
            sub = df_verif.copy()

        if sub.empty:
            return ""

        names = sub[cols.col_nombre].dropna().astype(str).str.strip()
        names = names[
            names.ne("")
            & ~names.str.lower().isin(["general", "nan", "none", "null", "—"])
        ]
        if names.empty:
            return ""

        # Modo por frecuencia: representa mejor quién estuvo cortando en el periodo.
        try:
            return str(names.mode().iloc[0]).strip()
        except Exception:
            return str(names.iloc[-1]).strip()
    except Exception:
        return ""

def _build_thb_export_workbook(day_exports: list[dict], period: str) -> io.BytesIO:
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side, PatternFill

    template_path = _get_thb_export_template_path()
    base_wb = load_workbook(template_path)

    if "Por hora" not in base_wb.sheetnames:
        raise RuntimeError("El modelo OEE debe tener una hoja llamada 'Por hora'.")

    # hoja maestra intacta
    master_ws = base_wb["Por hora"]

    # crear TODAS las hojas como copia de la maestra
    sheets_to_fill = []
    for item in day_exports:
        ws = base_wb.copy_worksheet(master_ws)
        ws.title = _safe_xlsx_sheet_title(item["sheet_title"])
        sheets_to_fill.append((ws, item))

    def _clear_row_values(ws, row_idx: int, col_start: int = 1, col_end: int = 22):
        for c in range(col_start, col_end + 1):
            ws.cell(row_idx, c).value = None

    def _clear_rows_values(ws, row_start: int, row_end: int, col_start: int = 1, col_end: int = 22):
        for r in range(row_start, row_end + 1):
            for c in range(col_start, col_end + 1):
                ws.cell(r, c).value = None

    def _clear_rows_borders(ws, row_start: int, row_end: int, col_start: int = 1, col_end: int = 22):
        empty_border = Border()
        for r in range(row_start, row_end + 1):
            for c in range(col_start, col_end + 1):
                ws.cell(r, c).border = empty_border

    fill_gray = PatternFill(fill_type="solid", fgColor="D9D9D9")
    fill_white = PatternFill(fill_type="solid", fgColor="FFFFFF")

    for ws, item in sheets_to_fill:
        rows = item.get("rows", []) or []

        # OEE esperado específico para exportación de crimpado.
        # Si no llega, se conserva el valor que tenga la plantilla en G2.
        try:
            oee_expected = item.get("oee_expected", None)
            if oee_expected is not None:
                ws["G2"] = float(oee_expected)
        except Exception:
            pass

        body_start = 7
        body_end = 38
        total_row_template = 40

        # activar filtros en fila 6
        ws.auto_filter.ref = f"A6:V6"

        body_end, _ = _ensure_thb_template_capacity(
            ws,
            needed_rows=len(rows),
            body_start=body_start,
            body_end=body_end,
            total_row=total_row_template,
        )

        _clear_thb_template_body(ws, body_start, body_end)
        _clear_rows_values(ws, body_start, max(ws.max_row, body_end + 10), 1, 22)
        _clear_rows_borders(ws, body_start, max(ws.max_row, body_end + 10), 1, 22)

        # escribir filas
        for idx, row in enumerate(rows, start=body_start):
            ws.cell(idx, 1).value = row["fecha"]
            ws.cell(idx, 2).value = row.get("referencia", "") or ""
            ws.cell(idx, 3).value = row.get("equipo", "THB") or "THB"
            ws.cell(idx, 4).value = row.get("operario", "") or ""
            ws.cell(idx, 5).value = row["inicio"]
            ws.cell(idx, 6).value = row["fin"]

            # Producción Planeada = OEE esperado * VN * tiempo base * factor de meta
            # Unión M1 usa TW (columna M) para que las paradas TX reduzcan la meta;
            # las demás máquinas conservan TP (columna K).
            try:
                meta_factor = float(row.get("meta_factor", item.get("meta_factor", 1.0)) or 1.0)
            except Exception:
                meta_factor = 1.0

            equipo_txt = str(row.get("equipo", "") or "").upper().replace(" ", "")
            plan_time_col = "M" if ("UNION" in equipo_txt and "M1" in equipo_txt) else "K"
            ws.cell(idx, 7).value = f'=IFERROR($G$2*O{idx}*{plan_time_col}{idx}*{meta_factor},"")'

            ws.cell(idx, 8).value = row["pn"]
            ws.cell(idx, 9).value = row["rx"]
            ws.cell(idx, 10).value = f'=IFERROR(H{idx}-I{idx},"")'
            ws.cell(idx, 11).value = row["tp_h"]
            ws.cell(idx, 12).value = row["tx_h"]
            ws.cell(idx, 13).value = f'=IFERROR(K{idx}-L{idx},"")'

            if row["tc_h_uni"] in (None, 0):
                ws.cell(idx, 14).value = None
            else:
                ws.cell(idx, 14).value = row["tc_h_uni"]

            ws.cell(idx, 15).value = f'=IFERROR(1/N{idx},"")'
            ws.cell(idx, 16).value = f'=IF(M{idx}=0,"",IFERROR(IF($P$2=1,H{idx}/O{idx},H{idx}*N{idx}),""))'
            ws.cell(idx, 17).value = f'=IFERROR(IF($Q$2=1,J{idx}/O{idx},J{idx}*N{idx}),"")'
            ws.cell(idx, 18).value = f'=IFERROR(M{idx}/K{idx},"")'
            ws.cell(idx, 19).value = f'=IFERROR(P{idx}/M{idx},"")'
            ws.cell(idx, 20).value = f'=IFERROR(Q{idx}/P{idx},"")'
            ws.cell(idx, 21).value = f'=IFERROR(Q{idx}/K{idx},"")'
            ws.cell(idx, 22).value = f'=IFERROR(H{idx}/G{idx},"")'

            ws.cell(idx, 1).number_format = "dd/mm/yyyy"
            ws.cell(idx, 5).number_format = "hh:mm"
            ws.cell(idx, 6).number_format = "hh:mm"
            ws.cell(idx, 11).number_format = "0.00"
            ws.cell(idx, 12).number_format = "0.00"
            ws.cell(idx, 14).number_format = "0.0000"

            # franjas alternadas
            row_fill = fill_gray if ((idx - body_start) % 2 == 0) else fill_white
            for col in range(1, 23):
                ws.cell(idx, col).fill = row_fill

        # TOTAL con una fila en blanco antes
        last_data_row = body_start + len(rows) - 1 if rows else body_start - 1
        blank_row = max(body_start, last_data_row + 1)
        total_row = blank_row + 1

        _clear_row_values(ws, blank_row, 1, 22)
        _clear_row_values(ws, total_row, 1, 22)

        _write_thb_template_totals(ws, body_start, max(last_data_row, body_start), total_row)

        # quitar palabra TOTAL
        ws.cell(total_row, 2).value = None

        ws.cell(total_row, 11).number_format = "0.00"
        ws.cell(total_row, 12).number_format = "0.00"
        ws.cell(total_row, 14).number_format = "0.0000"

        # línea del total
        thin_black = Side(style="thin", color="000000")
        for c in range(7, 23):
            ws.cell(total_row, c).border = Border(top=thin_black, bottom=thin_black)

        # ✅ UNIÓN 2: dejar visibles solo PB, RX (defectos), TP, TW y VN.
        # Se ocultan las demás columnas de indicadores (no se borran).
        # Visibles: A-F (contexto) + I(RX) J(PB) K(TP) M(TW) O(VN).
        # Ocultas:  G(PP) H(PN) L(TX) N(TC) P(TCN) Q(TCE) R(ID) S(IEO) T(IC) U(OEE) V(PA).
        if item.get("hide_extra_cols"):
            from openpyxl.utils import get_column_letter
            cols_ocultar = [7, 8, 12, 14, 16, 17, 18, 19, 20, 21, 22]
            for c in cols_ocultar:
                ws.column_dimensions[get_column_letter(c)].hidden = True

    # borrar hoja maestra original
    base_wb.remove(master_ws)

    if base_wb.worksheets:
        base_wb.active = 0

    bio = io.BytesIO()
    base_wb.save(bio)
    bio.seek(0)
    return bio


@app.get("/export_thb_excel")
def export_thb_excel():
    try:
        filename = (request.args.get("filename") or "").strip()
        machine = (request.args.get("machine") or "THB").strip().upper()
        period = (request.args.get("period") or "day").strip().lower()
        period_value = (request.args.get("period_value") or "").strip()
        operator = (request.args.get("operator") or "").strip()

        if machine != "THB":
            return jsonify({"ok": False, "error": "Esta exportación solo aplica para THB."}), 400
        if not filename:
            return jsonify({"ok": False, "error": "Falta filename."}), 400
        if not period_value:
            return jsonify({"ok": False, "error": "Falta period_value."}), 400

        df_verif, df_paradas = _load_thb_data_for_export(filename)
        days = _thb_days_for_period_export(df_verif, period, period_value, operator=operator)

        if not days:
            return jsonify({"ok": False, "error": "No se encontraron días con datos para exportar."}), 404

        day_exports = []
        for day in days:
            result_day = analyze_thb(
                df_verif,
                period="day",
                period_value=day.strftime("%Y-%m-%d"),
                machine="THB",
                operator=operator,
                time_start="",
                time_end="",
                df_paradas=df_paradas,
            )

            if not isinstance(result_day, dict) or result_day.get("status") != "ok":
                continue

            rows = _thb_export_rows_for_day(
                df_verif=df_verif,
                day=day,
                buckets=result_day.get("kpis", {}).get("hourly_buckets", []),
                operator=operator,
                df_paradas=df_paradas,
            )
            if not rows:
                continue

            day_exports.append({
                "sheet_title": day.strftime("%Y-%m-%d"),
                "rows": rows,
            })

        if not day_exports:
            return jsonify({"ok": False, "error": "No hubo datos exportables para los días seleccionados."}), 404

        period_tag = {
            "day": period_value,
            "week": period_value.replace("/", "-"),
            "month": period_value,
        }.get(period, period_value)

        download_name = f"THB_{period}_{period_tag}.xlsx".replace("/", "-").replace("\\", "-")
        payload = _build_thb_export_workbook(day_exports, period)

        resp = send_file(
            payload,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
        return resp

    except Exception as e:
        app.logger.exception("export_thb_excel failed")
        return jsonify({"ok": False, "error": f"Error exportando Excel THB: {e}"}), 500


@app.get("/export_hp_excel")
def export_hp_excel():
    try:
        filename = (request.args.get("filename") or "").strip()
        machine = (request.args.get("machine") or "HP").strip().upper()
        period = (request.args.get("period") or "day").strip().lower()
        period_value = (request.args.get("period_value") or "").strip()
        operator = (request.args.get("operator") or "").strip()
        time_start = (request.args.get("time_start") or "").strip()
        time_end = (request.args.get("time_end") or "").strip()
        time_apply_raw = (request.args.get("time_apply") or "").strip().lower()
        time_apply = time_apply_raw in ("1", "true", "yes", "si", "sí", "on")

        if machine != "HP":
            return jsonify({"ok": False, "error": "Esta exportación solo aplica para HP."}), 400
        if not filename:
            return jsonify({"ok": False, "error": "Falta filename."}), 400
        if not period_value:
            return jsonify({"ok": False, "error": "Falta period_value."}), 400

        df_corte, df_verif, df_paradas, df_duracion = _load_hp_data_for_export(filename)
        days = _hp_days_for_period_export(df_corte, period, period_value)

        if not days:
            return jsonify({"ok": False, "error": "No se encontraron días con datos de Corte para exportar."}), 404

        day_exports = _build_hp_export_day_exports(
            df_corte=df_corte,
            df_verif=df_verif,
            df_paradas=df_paradas,
            df_duracion=df_duracion,
            period=period,
            period_value=period_value,
            operator=operator,
            time_start=time_start,
            time_end=time_end,
            time_apply=time_apply,
        )

        if not day_exports:
            return jsonify({"ok": False, "error": "No hubo datos exportables para los días seleccionados."}), 404

        period_tag = {
            "day": period_value,
            "week": period_value.replace("/", "-"),
            "month": period_value,
        }.get(period, period_value)

        download_name = f"HP_{period}_{period_tag}.xlsx".replace("/", "-").replace("\\", "-")
        payload = _build_thb_export_workbook(day_exports, period)

        resp = send_file(
            payload,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
        return resp

    except Exception as e:
        app.logger.exception("export_hp_excel failed")
        return jsonify({"ok": False, "error": f"Error exportando Excel HP: {e}"}), 500


@app.get("/export_crimpado_excel")
def export_crimpado_excel():
    try:
        filename = (request.args.get("filename") or "").strip()
        machine = (request.args.get("machine") or "APLICACION").strip().upper()
        subcat = (request.args.get("subcat") or "").strip().lower()
        machine_id = (request.args.get("machine_id") or "").strip()
        period = (request.args.get("period") or "day").strip().lower()
        period_value = (request.args.get("period_value") or "").strip()
        operator = (request.args.get("operator") or "").strip()

        if machine not in ("APLICACION", "UNION"):
            return jsonify({"ok": False, "error": "Esta exportación solo aplica para máquinas de crimpado."}), 400
        if not filename:
            return jsonify({"ok": False, "error": "Falta filename."}), 400
        if not period_value:
            return jsonify({"ok": False, "error": "Falta period_value."}), 400

        # Aplicación/Unión trabaja normalmente por Google Sheets con este token.
        if filename == GSHEET_APP_TOKEN:
            path = GSHEET_APP_TOKEN
        else:
            path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
            if not os.path.exists(path):
                return jsonify({"ok": False, "error": f"No existe el archivo: {filename}"}), 400

        op2 = (operator or "").strip()
        if op2.lower() == "general":
            op2 = ""

        day_exports = get_crimpado_export_day_exports(
            path=path,
            period=period,
            period_value=period_value,
            machine=machine,
            subcat=subcat,
            operator=op2,
            machine_id=machine_id,
        )

        if not day_exports:
            return jsonify({"ok": False, "error": "No hubo datos exportables para la selección actual."}), 404

        period_tag = {
            "day": period_value,
            "week": period_value.replace("/", "-"),
            "month": period_value,
        }.get(period, period_value)

        machine_tag = f"{machine}_M{machine_id}" if machine_id else machine
        download_name = f"Crimpado_{machine_tag}_{period}_{period_tag}.xlsx".replace("/", "-").replace("\\", "-")

        payload = _build_thb_export_workbook(day_exports, period)

        resp = send_file(
            payload,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
        return resp

    except Exception as e:
        app.logger.exception("export_crimpado_excel failed")
        return jsonify({"ok": False, "error": f"Error exportando Excel Crimpado: {e}"}), 500


@app.post("/run_analysis_api")
def run_analysis_api():
    import os
    import time
    import pandas as pd
    from flask import request, jsonify

    def _nocache_json(payload, status=200):
        resp = jsonify(payload)
        resp.status_code = status
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
        return resp

    def _normalize_period_value(period: str, pv: str) -> str:
        pv = (pv or "").strip()
        if not pv:
            return ""

        # normaliza separadores
        if "/" in pv:
            pv = pv.replace("/", "-")

        # day: soporta YYYY-MM-DD y DD-MM-YYYY
        if period == "day":
            # YYYY-MM-DD
            if len(pv) == 10 and pv[4] == "-" and pv[7] == "-":
                return pv

            # DD-MM-YYYY -> YYYY-MM-DD
            if len(pv) == 10 and pv[2] == "-" and pv[5] == "-":
                dd = pv[0:2]
                mm = pv[3:5]
                yy = pv[6:10]
                if dd.isdigit() and mm.isdigit() and yy.isdigit():
                    return f"{yy}-{mm}-{dd}"

        # week/month: solo normaliza '/'
        if period in ("week", "month"):
            return pv

        return pv

    # =========================================================
    # ✅ Helpers locales
    # =========================================================
    def _find_col_local(df: pd.DataFrame, key: str):
        """Busca columna por contains (case-insensitive)."""
        k = (key or "").strip().lower()
        if df is None or df.empty:
            return None
        for c in df.columns:
            if k in str(c).strip().lower():
                return c
        return None

    def _build_dt_from_fecha_hora(df: pd.DataFrame, c_fecha: str, c_hora: str) -> pd.Series:
        """Construye datetime robusto desde Fecha + Hora."""
        if df is None or df.empty or not c_fecha or not c_hora:
            return pd.Series([], dtype="datetime64[ns]")

        # 1) intenta "Fecha Hora" directo
        s = (df[c_fecha].astype(str).fillna("").str.strip() + " " +
             df[c_hora].astype(str).fillna("").str.strip()).str.strip()
        dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
        if dt.notna().any():
            return dt

        # 2) intenta Fecha como date + Hora como timedelta
        d = pd.to_datetime(df[c_fecha], errors="coerce", dayfirst=True)
        td = pd.to_timedelta(df[c_hora].astype(str), errors="coerce")
        out = pd.to_datetime(d.dt.normalize() + td, errors="coerce")
        return out

    def _mask_by_period(dt: pd.Series, period: str, pv: str) -> pd.Series:
        """Retorna máscara booleana para day/week/month."""
        if dt is None or len(dt) == 0:
            return pd.Series([], dtype=bool)

        period = (period or "day").lower().strip()
        pv = (pv or "").strip()

        # si no hay period_value, no filtra
        if not pv:
            return dt.notna()

        if period == "day":
            target = pd.to_datetime(pv, errors="coerce")
            if pd.isna(target):
                return dt.notna()
            return dt.dt.date == target.date()

        if period == "month":
            # pv esperado: YYYY-MM (o algo que empiece con YYYY-MM)
            m = pv[:7]
            try:
                y = int(m[0:4])
                mo = int(m[5:7])
            except Exception:
                return dt.notna()
            return (dt.dt.year == y) & (dt.dt.month == mo)

        if period == "week":
            # pv esperado típico: YYYY-W##  (por ejemplo 2026-W05)
            import re
            m = re.search(r"(\d{4})\s*-\s*W(\d{1,2})", pv)
            if not m:
                m = re.search(r"(\d{4})\s*W(\d{1,2})", pv)
            if not m:
                return dt.notna()

            y = int(m.group(1))
            w = int(m.group(2))
            iso = dt.dt.isocalendar()
            return (iso["year"] == y) & (iso["week"] == w)

        return dt.notna()

    def _mask_by_time(dt: pd.Series, time_start: str, time_end: str) -> pd.Series:
        """Filtra por hora (HH:MM) usando datetime."""
        if dt is None or len(dt) == 0:
            return pd.Series([], dtype=bool)

        try:
            ts = pd.to_datetime(time_start, format="%H:%M", errors="coerce")
            te = pd.to_datetime(time_end, format="%H:%M", errors="coerce")
            if pd.isna(ts) or pd.isna(te):
                return dt.notna()
            t0 = ts.time()
            t1 = te.time()
            tt = dt.dt.time
            # caso normal
            if t0 <= t1:
                return dt.notna() & (tt >= t0) & (tt <= t1)
            # caso cruza medianoche
            return dt.notna() & ((tt >= t0) | (tt <= t1))
        except Exception:
            return dt.notna()

    # =========================================================
    # ✅ Helpers locales SOLO para Ciclo/Aplicadores (tal cual lo tenías)
    # =========================================================
    def _load_all_aplicacion_4maq(ttl_s: int | None = None) -> pd.DataFrame:
        ttl = int(getattr(config, "GSHEET_TTL_S", 10) or 10) if ttl_s is None else int(ttl_s)

        sources = [
            ("M1", getattr(config, "GSHEET_ID", None), getattr(config, "GSHEET_GID_APLICACION", None)),
            ("M2", getattr(config, "GSHEET_ID2", None), getattr(config, "GSHEET_GID_APLICACION2", None)),
            ("M3", getattr(config, "GSHEET_ID3", None), getattr(config, "GSHEET_GID_APLICACION3", None)),
            ("M4", getattr(config, "GSHEET_ID4", None), getattr(config, "GSHEET_GID_APLICACION4", None)),
        ]

        frames = []
        for tag, sid, gid in sources:
            if not sid or gid is None:
                continue
            df = load_gsheet_tab_csv_smart(str(sid), int(gid), ttl_s=ttl).copy()
            df["__src_machine"] = tag
            frames.append(df)

        if not frames:
            return pd.DataFrame()

        return pd.concat(frames, ignore_index=True)

    def _build_dt_series(df: pd.DataFrame) -> pd.Series:
        if df is None or df.empty:
            return pd.Series([], dtype="datetime64[ns]")

        c_fecha = _find_col_local(df, "fecha")
        c_hora = _find_col_local(df, "hora")

        if c_fecha and c_hora:
            s = (df[c_fecha].astype(str).fillna("").str.strip() + " " +
                 df[c_hora].astype(str).fillna("").str.strip()).str.strip()
            dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
            if dt.notna().any():
                return dt

        if c_fecha:
            dt = pd.to_datetime(df[c_fecha], errors="coerce", dayfirst=True)
            if dt.notna().any():
                return dt

        for guess in df.columns:
            g = str(guess).strip().lower()
            if "fecha" in g or "time" in g or "timestamp" in g:
                dt = pd.to_datetime(df[guess], errors="coerce", dayfirst=True)
                if dt.notna().any():
                    return dt

        return pd.Series([pd.NaT] * len(df), dtype="datetime64[ns]")

    def _load_union_pulsos_4books(ttl_s: int | None = None) -> pd.DataFrame:
        return _load_union_pulsos_4books_core(ttl_s=ttl_s)

    # =========================================================
    # ✅ NUEVO: Continuidad (Buses/Motos) por Google Sheets
    # =========================================================
    def _analyze_continuidad(df_hora: pd.DataFrame, df_dia: pd.DataFrame, period: str, pv: str, machine: str) -> dict:
        """
        Continuidad con salida visual tipo THB.

        Datos disponibles:
          - df_hora: Resumen Hora (Fecha, Hora, Referencia, Cantidad por referencia, Total por hora)
          - df_dia : Resumen Día x Ref (Fecha, Referencia, Cantidad del día)

        Regla solicitada:
          - Donde THB muestra "Metros Cortados", Continuidad muestra "Producción Total".
          - No se incluyen: Producción Buena, Producción con Defectos,
            Tiempo De Corte, Metros Extras, Metros Planeados.
          - Indicadores sin datos suficientes quedan en 0.
        """

        def _fmt_int_dot(n: Any) -> str:
            try:
                return f"{int(round(float(n))):,}".replace(",", ".")
            except Exception:
                return "0"

        def _fmt_hhmm_local(seconds: Any) -> str:
            try:
                sec = int(round(float(seconds)))
            except Exception:
                sec = 0
            sec = max(0, sec)
            hh = sec // 3600
            mm = (sec % 3600) // 60
            return f"{hh:02d}:{mm:02d}"

        def _zero_oee_ui() -> str:
            return (
                "0.0%||"
                "Indice de Eficiencia Operacional: 0.0% | "
                "Indice de Disponibilidad: 0.0% | "
                "Indice de Calidad: 0.0%"
            )

        def _empty_oee_chart(labels: list[str]) -> dict:
            labels = [str(x) for x in (labels or [])]
            return {
                "labels": labels,
                "oee": [0.0] * len(labels),
                "operacional": [0.0] * len(labels),
                "disponibilidad": [0.0] * len(labels),
                "calidad": [0.0] * len(labels),
            }

        if df_hora is None or df_hora.empty:
            return {"status": "error", "message": "Continuidad: hoja 'Resumen Hora' sin datos."}

        c_fecha = _find_col_local(df_hora, "fecha")
        c_hora = _find_col_local(df_hora, "hora")
        c_ref = _find_col_local(df_hora, "referencia")
        c_qty_ref = _find_col_local(df_hora, "cantidad")  # "Cantidad por referencia"
        c_total_h = _find_col_local(df_hora, "total por hora")  # exacto
        if not c_total_h:
            c_total_h = _find_col_local(df_hora, "total")  # fallback

        if not c_fecha or not c_hora or not c_ref:
            return {"status": "error",
                    "message": "Continuidad: faltan columnas (Fecha/Hora/Referencia) en Resumen Hora."}

        dt = _build_dt_from_fecha_hora(df_hora, c_fecha, c_hora)
        m = _mask_by_period(dt, period, pv)
        sub = df_hora.loc[m].copy()
        dt_sub = pd.to_datetime(dt.loc[m], errors="coerce") if hasattr(dt, "loc") else pd.to_datetime([],
                                                                                                      errors="coerce")

        if sub.empty:
            return {
                "status": "ok",
                "analysis": "continuidad",
                "title": "Análisis Continuidad",
                "machine": machine,
                "period": period,
                "period_value": pv,
                "rows_total": 0,
                "kpis_ui": {
                    "OEE": _zero_oee_ui(),
                    "Cumplimiento del plan": "0.0%",
                    "Producción Total": "0",
                    "Tiempo Pagado": "00:00",
                    "Tiempos Perdidos": "00:00||105: 00:00 | No registrado: 00:00 | 0.0%",
                    "Tiempo Trabajado": "00:00 (0.0%)",
                    "Referencias únicas": "0",
                    "Promedio por hora": "0.00",
                },
                "kpis": {"hourly_buckets": [], "other_causes": []},
                "terminal_usage": {"labels": [], "manual": [], "carrete": [], "otros": [], "total": []},
                "top_refs_day": [],
                "pareto": {"labels": [], "durations_sec": [], "cum_pct": []},
                "oee_chart": _empty_oee_chart([]),
            }

        # numéricos
        if c_qty_ref:
            sub[c_qty_ref] = pd.to_numeric(sub[c_qty_ref], errors="coerce").fillna(0)
        if c_total_h:
            sub[c_total_h] = pd.to_numeric(sub[c_total_h], errors="coerce").fillna(0)

        sub["__dt_cont"] = pd.to_datetime(dt_sub, errors="coerce").values
        sub = sub.loc[sub["__dt_cont"].notna()].copy()

        if sub.empty:
            horas_reg = 0
            total_pruebas = 0
        else:
            sub["__date_cont"] = sub["__dt_cont"].dt.normalize()
            sub["__hour_cont"] = sub["__dt_cont"].dt.hour.astype(int)

            # Total pruebas: si existe "Total por hora", tomar máximo por fecha/hora
            # para no duplicar el mismo total repetido por cada referencia.
            if c_total_h:
                total_por_hora = sub.groupby(["__date_cont", "__hour_cont"], dropna=False)[c_total_h].max()
                total_pruebas = int(total_por_hora.sum())
                horas_reg = int(total_por_hora.shape[0])
            elif c_qty_ref:
                total_por_hora = sub.groupby(["__date_cont", "__hour_cont"], dropna=False)[c_qty_ref].sum()
                total_pruebas = int(total_por_hora.sum())
                horas_reg = int(total_por_hora.shape[0])
            else:
                total_por_hora = sub.groupby(["__date_cont", "__hour_cont"], dropna=False).size()
                total_pruebas = int(total_por_hora.sum())
                horas_reg = int(total_por_hora.shape[0])

        refs_unicas = int(sub[c_ref].astype(str).str.strip().replace(
            {"": pd.NA, "nan": pd.NA}).dropna().nunique()) if c_ref in sub.columns and not sub.empty else 0
        promedio_hora = (total_pruebas / horas_reg) if horas_reg > 0 else 0.0

        # En Continuidad no hay modelo nominal ni defectos en estas hojas.
        # Por eso OEE, cumplimiento y sus componentes quedan en cero.
        tiempo_pagado_s = int(horas_reg * 3600)
        tiempo_trabajado_s = int(horas_reg * 3600) if total_pruebas > 0 else 0
        pct_trab = (tiempo_trabajado_s / tiempo_pagado_s * 100.0) if tiempo_pagado_s > 0 else 0.0

        kpis_ui = {
            "OEE": _zero_oee_ui(),
            "Cumplimiento del plan": "0.0%",
            "Producción Total": _fmt_int_dot(total_pruebas),

            "Tiempo Pagado": _fmt_hhmm_local(tiempo_pagado_s),
            "Tiempos Perdidos": "00:00||105: 00:00 | No registrado: 00:00 | 0.0%",
            "Tiempo Trabajado": f"{_fmt_hhmm_local(tiempo_trabajado_s)} ({pct_trab:.1f}%)",

            "Referencias únicas": _fmt_int_dot(refs_unicas),
            "Promedio por hora": f"{promedio_hora:.2f}",
        }

        # Referencias más usadas (se reutiliza terminal_usage para pintar tarjetas)
        if c_qty_ref and c_qty_ref in sub.columns:
            grp_ref = sub.groupby(c_ref, dropna=False)[c_qty_ref].sum().sort_values(ascending=False).head(12)
        else:
            grp_ref = sub.groupby(c_ref, dropna=False).size().sort_values(ascending=False).head(12)

        labels = [str(x) for x in grp_ref.index.tolist()]
        totals = [int(x) for x in grp_ref.values.tolist()]

        # Top referencias desde Resumen Día x Ref, si existe.
        top_refs = []
        if df_dia is not None and not df_dia.empty:
            cd_fecha = _find_col_local(df_dia, "fecha")
            cd_ref = _find_col_local(df_dia, "referencia")
            cd_qty = _find_col_local(df_dia, "cantidad")
            if cd_fecha and cd_ref and cd_qty:
                dd = pd.to_datetime(df_dia[cd_fecha], errors="coerce", dayfirst=True).dt.normalize()
                m2 = _mask_by_period(dd, period, pv)
                dsub = df_dia.loc[m2].copy()
                dsub[cd_qty] = pd.to_numeric(dsub[cd_qty], errors="coerce").fillna(0)
                grp = dsub.groupby(cd_ref, dropna=False)[cd_qty].sum().sort_values(ascending=False).head(10)
                top_refs = [{"label": str(k), "value": int(v)} for k, v in grp.items()]

        # Productividad por hora: solo para vista por día.
        hourly_buckets = []
        if str(period or "").strip().lower() == "day" and not sub.empty:
            if c_total_h:
                hour_totals = sub.groupby("__hour_cont", dropna=False)[c_total_h].max().sort_index()
            elif c_qty_ref:
                hour_totals = sub.groupby("__hour_cont", dropna=False)[c_qty_ref].sum().sort_index()
            else:
                hour_totals = sub.groupby("__hour_cont", dropna=False).size().sort_index()

            for h, val in hour_totals.items():
                try:
                    h_int = int(h)
                except Exception:
                    continue
                qty = int(round(float(val) if pd.notna(val) else 0))
                hourly_buckets.append({
                    "hour": f"{h_int:02d}:00",
                    "hourLabel": f"{h_int:02d}:00",
                    "cut": qty,
                    "meters": 0,
                    "prodSec": 3600 if qty > 0 else 0,
                    "deadSec": 0,
                    "otherDeadSec": 0,
                    "mealSec": 0,
                    "capacitySec": 3600,
                    "other_causes": [],
                    "len_mix": None,
                })

        # Gráfico OEE: mismo formato THB, valores en cero por falta de modelo nominal/calidad.
        chart_labels: list[str] = []
        if str(period or "").strip().lower() == "day":
            chart_labels = [b["hourLabel"] for b in hourly_buckets]
        elif not sub.empty:
            by_date = sub.groupby("__date_cont", dropna=False).size().sort_index()
            chart_labels = [pd.Timestamp(d).strftime("%d/%m") for d in by_date.index if pd.notna(d)]

        return {
            "status": "ok",
            "analysis": "continuidad",
            "title": "Análisis Continuidad",
            "machine": machine,
            "period": period,
            "period_value": pv,
            "rows_total": int(len(sub)),
            "kpis_ui": kpis_ui,
            "kpis": {
                "hourly_buckets": hourly_buckets,
                "other_causes": [],
            },
            "terminal_usage": {
                "labels": labels,
                "manual": totals,
                "carrete": [0] * len(totals),
                "otros": [0] * len(totals),
                "total": totals,
            },
            "top_refs_day": top_refs,
            "pareto": {"labels": [], "durations_sec": [], "cum_pct": []},
            "oee_chart": _empty_oee_chart(chart_labels),
        }

    try:
        result = None  # ✅ evita UnboundLocalError (result siempre existe)

        filename = (request.form.get("filename", "") or "").strip()
        analysis_key = (request.form.get("analysis_key", "") or "").strip()
        machine = (request.form.get("machine", "") or "").strip().upper()
        subcat = (request.form.get("subcat", "") or "").strip().lower()
        machine_id = (request.form.get("machine_id") or "").strip()

        # ✅ Normalizar analysis_key (alias)
        analysis_key = analysis_key.strip().lower()
        if analysis_key in ("aplicacion_excel", "app", "aplicacion1"):
            analysis_key = "aplicacion"
        if analysis_key in ("ciclo", "cicloaplicadores", "ciclo_aplicador", "ciclo-aplicadores"):
            analysis_key = "ciclo_aplicadores"

        period = (request.form.get("period", "day") or "day").strip().lower()
        period_value = (request.form.get("period_value", "") or "").strip()
        period_value = _normalize_period_value(period, period_value)

        operator = (request.form.get("operator", "") or "").strip() or "General"

        time_apply = (request.form.get("time_apply", "0") or "0").strip()
        time_apply = time_apply in ("1", "true", "True", "YES", "yes", "on")

        raw_time_start = (request.form.get("time_start", "") or "").strip()
        raw_time_end = (request.form.get("time_end", "") or "").strip()

        if period in ("week", "month"):
            time_start = ""
            time_end = ""
            time_apply = False
        else:
            time_start = raw_time_start or "00:00"
            time_end = raw_time_end or "23:59"

        print(
            "[run_analysis_api]",
            "machine=", machine,
            "machine_id=", machine_id,
            "period=", period,
            "period_value=", period_value,
            "operator=", operator,
            "time_apply=", time_apply,
            "time=", time_start, "-", time_end
        )

        if not filename or not analysis_key:
            return _nocache_json({"ok": False, "error": "Faltan parámetros: filename o analysis_key."}, 400)

        # ✅ tokens de Google Sheets (no son archivo real)
        is_gsheet_app = (analysis_key == "aplicacion" and filename == GSHEET_APP_TOKEN)
        is_gsheet_hp = (analysis_key == "corte" and machine == "HP" and filename == GSHEET_HP_TOKEN)
        is_gsheet_thb = (analysis_key == "corte" and machine == "THB" and filename == GSHEET_THB_TOKEN)
        is_gsheet_ciclo = (analysis_key == "ciclo_aplicadores" and filename == GSHEET_APP_TOKEN)

        # ✅ NUEVO: continuidad tokens (buses/motos)
        is_gsheet_cont_buses = (analysis_key == "continuidad" and filename == GSHEET_CONT_BUSES_TOKEN)
        is_gsheet_cont_motos = (analysis_key == "continuidad" and filename == GSHEET_CONT_MOTOS_TOKEN)

        if is_gsheet_app:
            path = GSHEET_APP_TOKEN
        elif is_gsheet_hp:
            path = GSHEET_HP_TOKEN
        elif is_gsheet_thb:
            path = GSHEET_THB_TOKEN
        elif is_gsheet_ciclo:
            path = GSHEET_APP_TOKEN
        elif is_gsheet_cont_buses:
            path = GSHEET_CONT_BUSES_TOKEN
        elif is_gsheet_cont_motos:
            path = GSHEET_CONT_MOTOS_TOKEN
        else:
            path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
            if not os.path.exists(path):
                return _nocache_json({"ok": False, "error": f"No existe el archivo: {filename}"}, 400)

        # =========================================================
        # CORTE: THB / HP
        # =========================================================
        if analysis_key == "corte":
            if machine == "THB":
                cache_thb = THB_CACHE.get(filename)

                if filename == GSHEET_THB_TOKEN:
                    sheet_id = getattr(config, "THB_SHEET_ID", "") or ""
                    gid_verif = int(getattr(config, "THB_GID_VERIF", 0) or 0)
                    gid_par = int(getattr(config, "THB_GID_PARADAS", 0) or 0)
                    ttl = int(getattr(config, "GSHEET_TTL_S", 10) or 10)

                    if not sheet_id or not gid_verif or not gid_par:
                        return _nocache_json(
                            {"ok": False, "error": "Falta config THB en config.py (THB_SHEET_ID/THB_GID_*)."}, 400)

                    df_verif = load_gsheet_tab_csv_smart(sheet_id, gid_verif, ttl_s=ttl).copy()
                    df_paradas = load_gsheet_tab_csv_smart(sheet_id, gid_par, ttl_s=ttl).copy()
                    df_paradas = _normalize_gsheet_paradas_duration_to_seconds(df_paradas)
                    df_paradas = _hp_gsheet_fix_paradas_duration(df_paradas)

                    df_duracion = pd.DataFrame()

                    try:
                        if filename in THB_CACHE and isinstance(THB_CACHE[filename], dict):
                            THB_CACHE[filename]["df"] = df_verif.copy()
                            THB_CACHE[filename]["df_paradas"] = df_paradas.copy()
                        else:
                            THB_CACHE[filename] = {"df": df_verif.copy(), "df_paradas": df_paradas.copy()}
                    except Exception:
                        pass

                else:
                    if cache_thb and isinstance(cache_thb.get("df"), pd.DataFrame):
                        df_verif = cache_thb["df"].copy()
                        df_paradas = cache_thb.get("df_paradas", pd.DataFrame()).copy()
                    else:
                        try:
                            df_verif = _read_excel_smart(path, "Verificación de Medidas").copy()
                        except Exception:
                            df_verif = _read_excel_smart(path, "Verificacion de Medidas").copy()

                        df_paradas = _read_excel_smart(path, "Paradas").copy()

                        try:
                            if filename in THB_CACHE and isinstance(THB_CACHE[filename], dict):
                                THB_CACHE[filename]["df"] = df_verif.copy()
                                THB_CACHE[filename]["df_paradas"] = df_paradas.copy()
                            else:
                                THB_CACHE[filename] = {"df": df_verif.copy(), "df_paradas": df_paradas.copy()}
                        except Exception:
                            pass

                result = analyze_thb(
                    df_verif,
                    period=period,
                    period_value=period_value,
                    machine="THB",
                    operator=operator,
                    time_start=(time_start if time_apply else ""),
                    time_end=(time_end if time_apply else ""),
                    df_paradas=df_paradas,
                )

                # ✅ THB: el dashboard debe mostrar los mismos OEE/PA/TW/TX,
                # producción total y producción planeada que se calculan
                # en el Excel descargable "Descargar Excel THB".
                if isinstance(result, dict) and result.get("status") == "ok":
                    thb_oee_expected = _thb_export_oee_expected(default=0.70)
                    thb_day_exports = _build_thb_export_day_exports(
                        df_verif=df_verif,
                        df_paradas=df_paradas,
                        period=period,
                        period_value=period_value,
                        operator=operator,
                        base_result=result,
                    )
                    result = _thb_apply_excel_metrics_to_result(
                        result,
                        thb_day_exports,
                        oee_expected=thb_oee_expected,
                    )

                    thb_operator_title_name = _thb_productivity_operator_name_from_data(
                        df_verif,
                        period=period,
                        period_value=period_value,
                        operator=operator,
                    )
                    if thb_operator_title_name:
                        result["productivity_operator_name"] = thb_operator_title_name
                        result["current_operator_name"] = thb_operator_title_name

            elif machine == "HP":
                if filename == GSHEET_HP_TOKEN:
                    sheet_id = getattr(config, "HP_SHEET_ID", "") or ""
                    gid_corte = int(getattr(config, "HP_GID_CORTE", 0) or 0)
                    gid_verif = int(getattr(config, "HP_GID_VERIF", 0) or 0)
                    gid_par = int(getattr(config, "HP_GID_PARADAS", 0) or 0)
                    ttl = int(getattr(config, "GSHEET_TTL_S", 10) or 10)

                    if not sheet_id or not gid_corte or not gid_verif or not gid_par:
                        return _nocache_json({
                            "ok": False,
                            "error": "Falta config de HP en config.py (HP_SHEET_ID y HP_GID_*)."
                        }, 400)

                    df_corte = load_gsheet_tab_csv_smart(sheet_id, gid_corte, ttl_s=ttl).copy()
                    df_verif = load_gsheet_tab_csv_smart(sheet_id, gid_verif, ttl_s=ttl).copy()
                    df_paradas = load_gsheet_tab_csv_smart(sheet_id, gid_par, ttl_s=ttl).copy()

                    df_paradas = _normalize_gsheet_paradas_duration_to_seconds(df_paradas)
                    df_paradas = _hp_gsheet_fix_paradas_duration(df_paradas)

                    df_duracion = pd.DataFrame()

                else:
                    df_corte = _read_excel_smart(path, "Corte").copy()
                    try:
                        df_verif = _read_excel_smart(path, "Verificación de Medidas").copy()
                    except Exception:
                        df_verif = _read_excel_smart(path, "Verificacion de Medidas").copy()

                    df_paradas = _read_excel_smart(path, "Paradas").copy()

                    try:
                        df_duracion = pd.read_excel(path, sheet_name="Duración", header=None, engine="openpyxl")
                    except Exception:
                        try:
                            df_duracion = pd.read_excel(path, sheet_name="Duracion", header=None, engine="openpyxl")
                        except Exception:
                            df_duracion = pd.DataFrame()

                result = analyze_hp(
                    df_corte,
                    df_verif,
                    period=period,
                    period_value=period_value,
                    machine="HP",
                    operator=operator,
                    time_start=(time_start if time_apply else ""),
                    time_end=(time_end if time_apply else ""),
                    time_apply=time_apply,
                    df_paradas=df_paradas,
                    df_duracion=df_duracion,
                )

                # ✅ HP: el dashboard debe mostrar los mismos OEE/PA/TW/TX
                # que se calculan en el Excel descargable.
                if isinstance(result, dict) and result.get("status") == "ok":
                    hp_day_exports = _build_hp_export_day_exports(
                        df_corte=df_corte,
                        df_verif=df_verif,
                        df_paradas=df_paradas,
                        df_duracion=df_duracion,
                        period=period,
                        period_value=period_value,
                        operator=operator,
                        time_start=(time_start if time_apply else ""),
                        time_end=(time_end if time_apply else ""),
                        time_apply=time_apply,
                        base_result=result,
                    )
                    result = _hp_apply_excel_metrics_to_result(
                        result,
                        hp_day_exports,
                        oee_expected=0.70,
                    )

            else:
                return _nocache_json({"ok": False, "error": f"Máquina no soportada: {machine}"}, 400)

            if result is None:
                return _nocache_json({"ok": False, "error": "Error interno: result=None (no se ejecutó el análisis)."},
                                     500)
            if not isinstance(result, dict):
                return _nocache_json({"ok": False, "error": "El análisis no devolvió un dict válido."}, 500)
            if result.get("status") != "ok":
                return _nocache_json({"ok": False, "error": result.get("message", "Error ejecutando análisis.")}, 400)

            return _nocache_json({"ok": True, "result": result}, 200)

        # =========================================================
        # APLICACION / UNION  (NO TOCAR)
        # =========================================================
        elif analysis_key == "aplicacion":
            if machine not in ("APLICACION", "UNION"):
                if subcat in ("aplicacion", "union"):
                    machine = "APLICACION" if subcat == "aplicacion" else "UNION"
                else:
                    machine = "APLICACION"

            op2 = (operator or "").strip()
            if op2.lower() == "general":
                op2 = ""

            try:
                result = run_aplicacion_analysis(
                    path=path,
                    period=period,
                    period_value=period_value,
                    machine=machine,
                    subcat=subcat,
                    operator=op2,
                    time_start=(time_start if time_apply else ""),
                    time_end=(time_end if time_apply else ""),
                    machine_id=machine_id,
                )
            except TypeError:
                result = run_aplicacion_analysis(
                    path=path,
                    period=period,
                    period_value=period_value,
                    machine=machine,
                    subcat=subcat,
                    operator=op2,
                    time_start=(time_start if time_apply else ""),
                    time_end=(time_end if time_apply else ""),
                )

            if not isinstance(result, dict):
                return _nocache_json({"ok": False, "error": "El análisis APLICACION no devolvió un dict válido."}, 500)

            out = {"status": "ok", **result}
            return _nocache_json({"ok": True, "result": out}, 200)

        # =========================================================
        # ✅ CICLO / APLICADORES (UNION 4 libros)
        # =========================================================
        elif analysis_key == "ciclo_aplicadores":
            df = _load_union_pulsos_4books()

            if df.empty:
                return _nocache_json({"ok": False,
                                      "error": "No hay datos válidos en UNION (se requieren columnas Terminal y Pulsos en el resumen)."},
                                     400)

            grp = df.groupby("terminal", dropna=True)["pulsos"].sum().sort_values(ascending=False)

            labels = grp.index.astype(str).tolist()
            totals = [int(x) for x in grp.values.tolist()]

            result = {
                "status": "ok",
                "machine": "UNION",
                "period": "total",
                "period_value": "",
                "rows_total": int(len(df)),
                "kpis_ui": {
                    "Terminales únicos": str(int(len(labels))),
                    "Pulsos totales": str(int(grp.sum())),
                },
                "terminal_usage": {
                    "labels": labels,
                    "manual": totals,
                    "carrete": [0] * len(totals),
                    "otros": [0] * len(totals),
                    "total": totals,
                },
                "note": "UNION: suma la columna Pulsos por Terminal. Para Unión 1/2 usa UNION*_Resumen_Hora cuando el GID esté configurado.",
            }
            return _nocache_json({"ok": True, "result": result}, 200)

        # =========================================================
        # ✅ CONTINUIDAD (BUSES / MOTOS) por Google Sheets
        # =========================================================
        elif analysis_key == "continuidad":
            ttl = int(getattr(config, "GSHEET_TTL_S", 10) or 10)

            if filename == GSHEET_CONT_BUSES_TOKEN:
                sheet_id = getattr(config, "CONT_BUSES_SHEET_ID", "") or ""
                gid_hora = int(getattr(config, "CONT_BUSES_GID_RESUMEN_HORA", 0) or 0)
                gid_dia = int(getattr(config, "CONT_BUSES_GID_RESUMEN_DIA_REF", 0) or 0)
                machine = "BUSES"
            elif filename == GSHEET_CONT_MOTOS_TOKEN:
                sheet_id = getattr(config, "CONT_MOTOS_SHEET_ID", "") or ""
                gid_hora = int(getattr(config, "CONT_MOTOS_GID_RESUMEN_HORA", 0) or 0)
                gid_dia = int(getattr(config, "CONT_MOTOS_GID_RESUMEN_DIA_REF", 0) or 0)
                machine = "MOTOS"
            else:
                return _nocache_json(
                    {"ok": False, "error": "Continuidad requiere token de Google Sheets (buses/motos)."}, 400)

            if not sheet_id or gid_hora is None or gid_dia is None:
                return _nocache_json({"ok": False, "error": "Falta config Continuidad en config.py."}, 400)

            df_hora = load_gsheet_tab_csv_smart(sheet_id, gid_hora, ttl_s=ttl).copy()
            df_dia = load_gsheet_tab_csv_smart(sheet_id, gid_dia, ttl_s=ttl).copy()

            result = _analyze_continuidad(df_hora, df_dia, period=period, pv=period_value, machine=machine)

            if not isinstance(result, dict):
                return _nocache_json({"ok": False, "error": "El análisis Continuidad no devolvió un dict válido."}, 500)
            if result.get("status") != "ok":
                return _nocache_json({"ok": False, "error": result.get("message", "Error ejecutando continuidad.")},
                                     400)
            # ✅ COMPAT: expón KPIs en 2 nombres
            if "kpis_ui" in result and "kpis" not in result:
                result["kpis"] = result["kpis_ui"]

            return _nocache_json({"ok": True, "result": result}, 200)

        return _nocache_json({"ok": False, "error": f"analysis_key no soportado: {analysis_key}"}, 400)


    except Exception as e:
        return _nocache_json({"ok": False, "error": f"Error interno: {e}"}, 500)


@app.route("/run", methods=["POST"])
def run_analysis():
    machine = request.form.get("machine", "")
    if machine == "THB":
        # Si aún usas esta ruta, queda igual.
        flash("Esta ruta /run no se usa en el modo AJAX actual.", "error")
        return redirect(url_for("home"))

    flash("HP aún no implementado (estructura distinta).", "error")
    return redirect(url_for("home"))


# =========================
# MAIN
# =========================
if __name__ == "__main__":
    import webbrowser

    webbrowser.open("http://127.0.0.1:5000")
    app.run(debug=False, host="127.0.0.1", port=5000)